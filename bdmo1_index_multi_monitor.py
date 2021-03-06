# ‐*‐ coding: utf‐8 ‐*‐
"""
功能:
   1)指定几个域名,分关键词种类监控首页词数
   2)采集serp所有url,提取域名并统计各域名首页覆盖率
   3)采集了serp上的排名url特征srcid值
提示:
  1)相关网站.相关企业.智能小程序.其他人还在搜.热议聚合.资讯聚合.搜索智能聚合.视频全部算在内
    所以首页排名有可能大于10,未提取多卡片的样式
  2)serp上自然排名mu属性值为排名url,特殊样式mu为空或不存在,
    提取article里url,该url是baidu域名,二次访问才能获得真实url,本脚本直接取baidu链接
  3)2020xiaoqu_kwd_city_new.xlsx:sheet名为关键词种类,sheet第一列放关键词
结果:
    bdmo1_index_info.txt:各监控站点词的排名及url,如有2个url排名,只取第一个
    bdmo1_index_all.txt:serp所有url及样式特征,依此统计各域名首页覆盖率-单写脚本完成
    bdmo1_index.xlsx:自己站每类词首页词数
    bdmo1_index_domains.xlsx:各监控站点每类词的首页词数
    bdmo1_index_domains.txt:各监控站点每类词的首页词数
header头信息复制浏览器全部的请求头,Accept-Encoding留deflate
serp上请求头多测几个找到和手动搜索结果最一致的
"""

import requests
from pyquery import PyQuery as pq
import threading
import queue
import time
from urllib.parse import urlparse
from openpyxl import load_workbook
from openpyxl import Workbook
import time
import gc
import json
import random

# 计算最终结果
def get_result(file_path, result):
    for line in open(file_path, 'r', encoding='utf-8'):
        line = line.strip().split('\t')
        rank = line[2]
        group = line[3]
        domain = line[4]
        if rank != '无':
            result[domain][group]['首页'] += 1
        result[domain][group]['总词数'] += 1
    return result


# 写txt,所有监控域名的结果
def write_domains_txt(result_last):
    with open('{0}bdmo1_index_domains.txt'.format(today), 'w', encoding="utf-8") as f_res:
        f_res.write('{0}\t{1}\t{2}\t{3}\t{4}\n'.format('日期','域名','词类','首页词数','查询词数'))
        for now_domain,dict_value in result_last.items():
            for group, dict_index_all in dict_value.items():
                f_res.write('{0}\t{1}\t{2}\t'.format(today,now_domain,group))
                for key, value in dict_index_all.items():
                    f_res.write(str(value) + '\t')
                f_res.write('\n')


# 写excel
def write_myexcel(group_list, result_last, today,my_domain):
    wb = Workbook()
    wb_all = Workbook()
    # 创建sheet写表头
    for group in group_list:
        sheet_num = 0
        wb.create_sheet(u'{0}'.format(group), index=sheet_num)
        wb_all.create_sheet(u'{0}'.format(group), index=sheet_num)
        row_first = ['日期', '首页', '总词数']
        row_first2 = ['日期', '域名','首页', '总词数']
        # 写表头
        wb[group].append(row_first)
        wb_all[group].append(row_first2)
        sheet_num += 1
    # 写内容
    for domain, dict_value in result_last.items():
        if domain == my_domain:
            for group, dict_index_all in dict_value.items():
                # 写数据
                row_value = [today]
                for key,value in dict_index_all.items():
                    row_value.append(value)
                wb[u'{0}'.format(group)].append(row_value)

        for group, dict_index_all in dict_value.items():
            # 写数据
            row_value = [today,domain]
            for key, value in dict_index_all.items():
                row_value.append(value)
            wb_all[u'{0}'.format(group)].append(row_value)
    wb.save('{0}bdmo1_index.xlsx'.format(today))
    wb_all.save('{0}bdmo1_index_domains.xlsx'.format(today))

# 发js包-不用
# def request_js(url,my_header,retry=1):
#     try:
#         r = requests.get(url=url,headers=my_header,timeout=2)
#     except Exception as e:
#         print('获取源码失败',e)
#         time.sleep(6)
#         if retry > 0:
#             request_js(url,retry-1)
#     else:
#         pass

# header
def get_header():
    my_header = {
    'Accept':'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
'Accept-Encoding':'deflate',
'Accept-Language':'zh-CN,zh;q=0.9,en;q=0.8',
'Cache-Control':'max-age=0',
'Connection':'keep-alive',
'Cookie':'BAIDUID=37E3E4778B6D89A1959329D105A13A16:FG=1; H_WISE_SIDS=148078_149391_150723_149985_150077_147088_150085_148193_148867_148713_150745_147280_150647_150165_149587_150155_148754_147896_146575_148523_151033_127969_149908_146551_146453_149719_147353_146732_138425_149557_143667_131423_100806_147528_151148_107319_148185_149253_140367_144966_150341_147546_148868_110085; rsv_i=4810ma%2B%2BO%2BXmFDzVk5gkLRtJ2q7oESxO9pg1Hnlq51IETtzPgmCfYgm3%2BNjKTKvFOUTnyXIs%2B6hoOhi7xnNAtxxnibkX8sY; plus_lsv=e1339ee5f098ff6b; BDORZ=AE84CDB3A529C0F8A2B9DCDD1D18B695; plus_cv=1::m:49a3f4a6; bd_af=1; BDICON=10123156; delPer=0; BDORZ=SFH; wpr=0; MSA_WH=375_667; MSA_PHY_WH=750_1334; MSA_PBT=146; MSA_ZOOM=1056; SE_LAUNCH=5%3A26568472_15%3A26568472; ysm=10557|10557; BA_HECTOR=kj094; BDSVRBFE=Go; __bsi=7965975238557868905_00_7_R_R_7_0303_c02f_Y; COOKIE_SESSION=6_0_0_2_0_t3_8_2_1_0_0_2_7_1594108380%7C2%230_0_0_0_0_0_0_0_1594108374%7C1; FC_MODEL=0_0_0_0_2.39_0_0_0_0_0_2.92_0_2_8_0_3_0_0_1594108380%7C2%232.39_0_0_2_0_0_1594108380%7C2%230_ax_0_0_0_0_51_1594108380',
'Host':'m.baidu.com',
'Referer':'https://m.baidu.com/s?word=%E5%8C%97%E4%BA%AC%E7%A7%9F%E6%88%BF&ts=8373610&t_kt=0&ie=utf-8&rsv_iqid=3169232709&rsv_t=466fteG54%252FWirs4NDCgKpXRN1BgWcOiMV3UwnQiZRhplHhY24mR7&sa=ib&rsv_pq=3169232709&rsv_sug4=3694&inputT=2170&sugid=2988971856915527823&ss=100&tj=1',
'Sec-Fetch-Dest':'document',
'Sec-Fetch-Mode':'navigate',
'Sec-Fetch-Site':'same-origin',
'Sec-Fetch-User':'?1',
'Upgrade-Insecure-Requests':'1',
'User-Agent':'Mozilla/5.0 (iPhone; CPU iPhone OS 13_2_3 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/13.0.3 Mobile/15E148 Safari/604.1',
     }     
    return my_header



class bdmoIndexMonitor(threading.Thread):

    def __init__(self):
        threading.Thread.__init__(self)

    @staticmethod
    def read_excel(filepath):
        q = queue.Queue()
        group_list = []
        kwd_dict = {}
        wb_kwd = load_workbook(filepath)
        for sheet_obj in wb_kwd:
            sheet_name = sheet_obj.title
            group_list.append(sheet_name)
            kwd_dict[sheet_name]= []
            col_a = sheet_obj['B']
            for cell in col_a:
                kwd = (cell.value)
                # 加个判断吧
                if kwd:
                    q.put([sheet_name,kwd])
        return q, group_list

    # 初始化结果字典
    @staticmethod
    def result_init(group_list):
        result = {}
        for domain in domains:
            result[domain] = {}
            for group in group_list:
                result[domain][group] = {'首页':0,'总词数':0}
        print("结果字典init...")
        return result

    # 获取某词serp源码
    def get_html(self,url,my_header,retry=1):
        try:
            r = requests.get(url=url,headers=my_header,timeout=5)
        except Exception as e:
            print('获取源码失败',e)
            time.sleep(6)
            if retry > 0:
                self.get_html(url,my_header,retry-1)
        else:
            html = r.content.decode('utf-8',errors='ignore')  # 用r.text有时候识别错误
            # print(html)
            url = r.url  # 反爬会重定向,取定向后的地址
            return html,url

    # 获取某词的serp源码上包含排名url的div块
    def get_divs(self, html ,url):
        div_list = []
        doc = pq(html)
        title = doc('title').text()
        if '- 百度' in title and 'https://m.baidu.com/s?ie=utf-8' in url:
            try:
                div_list = doc('.c-result').items()
            except Exception as e:
                print('提取div块失败', e)
            else:
                pass
        else:
            print('源码异常------',title)
            time.sleep(360)
        return div_list

    # 提取排名的真实url
    def get_real_urls(self, div_list):
        real_urls_rank = []
        if div_list:
            for div in div_list:
                try:
                    data_log = div.attr('data-log')
                    data_log = json.loads(data_log.replace("'", '"')) # json字符串双引号
                    srcid = data_log['ensrcid'] if 'ensrcid' in data_log  else 'ensrcid' # 样式特征
                    rank_url = data_log['mu']  if 'mu' in data_log else None # mu可能为空或者不存在
                    rank = data_log['order']
                except Exception as e:
                    print('提取rank_url error',e)
                else:
                    if rank_url:
                        real_urls_rank.append((rank_url,rank,srcid))
                    # 如果mu为空或者不存在
                    else:
                        # 提取资讯聚合,图片聚合
                        article = div('.c-result-content article')
                        link = article.attr('rl-link-href')
                        # 提取热议聚合
                        if not link:
                            a = div('.c-result-content article header a')
                            data_log_ugc = a.attr('data-log')
                            data_log_ugc = json.loads(data_log_ugc.replace("'", '"')) # json字符串双引号
                            link = data_log_ugc['mu']  if 'mu' in data_log_ugc else None # mu可能为空或者不存在
                            link = 'https://m.baidu.com{0}'.format(link) if link != None else None
                            # 一般为卡片样式,链接太多,不提取了
                            if not link:
                                pass
                        real_urls_rank.append((link,rank,srcid))
        return real_urls_rank

    # 提取某url的域名部分
    def get_domain(self,real_url):
        domain = None
        try:
           res = urlparse(real_url)
        except Exception as e:
           print(e,'real_url:error')
        else:
           domain = res.netloc
        return domain

    # 获取某词serp源码首页排名所有域名
    def get_domains(self,real_url_list):
            domain_list = [self.get_domain(real_url) for real_url in real_url_list]
            # 一个词某域名多个url有排名,算一次
            domain_set = set(domain_list)
            domain_set = domain_set.remove(None) if None in domain_set else domain_set
            domain_str = ','.join(domain_set)
            return domain_str

    # 线程函数
    def run(self):
        # js_url = 'https://fclick.baidu.com/w.gif?baiduid=14E9731020ACEE14821E1A67DABB2862&asp_time=1581297830764&query={0}&queryUtf8={1}&searchid=a0bc28b872b56b7e&osid=1&bwsid=5&adt=0&adb=0&wst=146&top=0&wise=10&middle=0&bottom=0&adpos=t_0_0.00&pbt=146&yxh=0&zoom=1.0555555555555556&validHeight=521&initViewZone=w_1_0.00%3Aw_2_1.00&adsHeight=_w1%3A255_w2%3A255_w3%3A487_w4%3A228_w5%3A204_w6%3A165_w7%3A189_w8%3A255_w9%3A151_w10%3A103&adsCmatch=&availHeight=667&availWidth=375&winHeight=667&winWidth=375&action=init&model=%7B%22vt%22%3A%22w1%3A0%23w2%3A0%23w3%3A0%23w4%3A0%23w5%3A0%23w6%3A0%23w7%3A0%23w8%3A0%23w9%3A0%23w10%3A0%22%2C%22pt%22%3A%22%22%2C%22ext%22%3A%5B%5D%2C%22vsh%22%3A521%2C%22asid%22%3A%22%22%2C%22rd%22%3A1581297833317%7D&tag=ecom_wise_listen_n&rand=1581297833325.636'
        while 1:
            group_kwd = q.get()
            group,kwd = group_kwd
            print(group,kwd)
            try:
                url = "https://m.baidu.com/s?ie=utf-8&word={0}".format(kwd)
                # js_url = js_url.format(kwd,kwd)
                my_header = get_header()
                # request_js(js_url,my_header)
                html,now_url = self.get_html(url,my_header)
                divs_res = self.get_divs(html,now_url)
                # 源码ok再写入
                if divs_res:
                    real_urls_rank = self.get_real_urls(divs_res)
                    real_urls = []
                    for my_url,my_order,my_attr in real_urls_rank:
                        real_urls.append(my_url)
                        f_all.write('{0}\t{1}\t{2}\t{3}\t{4}\n'.format(kwd,my_url,my_order,my_attr,group))
                    f_all.flush()
                    domain_str = self.get_domains(real_urls)
                    # 目标站点是否出现
                    for domain in domains:
                        if domain not in domain_str:
                              f.write('{0}\t{1}\t{2}\t{3}\t{4}\n'.format(kwd, '无', '无', group,domain))
                        else:
                            for my_url,my_order,my_attr in real_urls_rank:
                                if domain in my_url:
                                    f.write('{0}\t{1}\t{2}\t{3}\t{4}\t{5}\n'.format(kwd,my_url,my_order,group,domain,my_attr))
                                    # print(my_url, my_order)
                                    break # 取第一个排名url
                f.flush()
            except Exception as e:
                print(e,'error')
                print(real_urls)
            finally:
                del kwd
                gc.collect()
                q.task_done()
                # exit()
                time.sleep(0.2)
                

if __name__ == "__main__":
    start = time.time()
    local_time = time.localtime()
    today = time.strftime('%Y%m%d',local_time)
    domains = ['5i5j.com','lianjia.com','anjuke.com','fang.com'] # 目标域名
    my_domain = '5i5j.com' # 自己域名
    q,group_list = bdmoIndexMonitor.read_excel('2020xiaoqu_kwd_city_new.xlsx')  # 关键词队列及分类
    result = bdmoIndexMonitor.result_init(group_list)  # 初始化结果
    all_num = q.qsize() # 总词数
    f = open('{0}bdmo1_index_info.txt'.format(today),'w',encoding="utf-8")
    f_all = open('{0}bdmo1_index_all.txt'.format(today),'w',encoding="utf-8")
    file_path = f.name
    # 设置线程数
    for i in list(range(1)):
        t = bdmoIndexMonitor()
        t.setDaemon(True)
        t.start()
    q.join()
    f.close()
    f_all.close()
    # 根据bdmo1_index_info.txt计算结果
    result_last = get_result(file_path,result)
    # 写入txt文件
    write_domains_txt(result_last)
    # 写入excel
    write_myexcel(group_list,result_last,today,my_domain)
    # 统计查询成功的词数
    with open(file_path,'r',encoding='utf-8') as fp:
         success = int(sum(1 for x in fp)/len(domains))
    end = time.time()
    print('关键词共{0}个,查询成功{1}个,耗时{2}min'.format(all_num,success,(end - start) / 60))

# ‐*‐ coding: utf‐8 ‐*‐
"""
功能:
   1)指定几个域名,分关键词种类监控首页词数
   2)采集serp所有url,提取域名并统计各域名首页覆盖率
   3)采集了serp上的排名url特征srcid值
提示:
  1)相关网站.相关企业.智能小程序.其他人还在搜.热议聚合.资讯聚合.搜索智能聚合.视频全部算在内
    所以首页排名有可能大于10
  2)serp上自然排名mu属性值为排名url,特殊样式mu为空或不存在,
    提取article里url,该url是baidu域名,二次访问才能获得真实url,本脚本直接取baidu链接
  3)2020kwd_url_core_city_unique.xlsx:sheet名为关键词种类,sheet第一列放关键词
结果:
    bdmo1_index_info.txt:各监控站点词的排名及url,如有2个url排名,只取第一个
    bdmo1_index_all.txt:serp所有url及样式特征,依此统计各域名首页覆盖率-单写脚本完成
    bdmo1_index.xlsx:自己站每类词首页词数
    bdmo1_index_domains.xlsx:各监控站点每类词的首页词数
    bdmo1_index_domains.txt:各监控站点每类词的首页词数
selenium驱动浏览器的方式 默认为无头模式,selenium不能支持上时间运行浏览器，代码采用有异常重启的方式持续
"""

from pyquery import PyQuery as pq
import threading
import queue
import time
import json
from urllib.parse import urlparse
from openpyxl import load_workbook
from openpyxl import Workbook
import time
import gc
import random
import requests
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service


requests.packages.urllib3.disable_warnings()

# 计算最终结果
def get_result(file_path, result):
    for line in open(file_path, 'r', encoding='utf-8'):
        line = line.strip().split('\t')
        rank = line[2]
        group = line[3]
        domain = line[4]
        if rank != '无':
            result[domain][group]['首页'] += 1
        result[domain][group]['总词数'] += 1
    return result


# 写txt,所有监控域名的结果
def write_domains_txt(result_last):
    with open('{0}bdmo1_index_domains.txt'.format(today), 'w', encoding="utf-8") as f_res:
        f_res.write('{0}\t{1}\t{2}\t{3}\t{4}\n'.format('日期','域名','词类','首页词数','查询词数'))
        for now_domain,dict_value in result_last.items():
            for group, dict_index_all in dict_value.items():
                f_res.write('{0}\t{1}\t{2}\t'.format(today,now_domain,group))
                for key, value in dict_index_all.items():
                    f_res.write(str(value) + '\t')
                f_res.write('\n')


# 写excel
def write_myexcel(group_list, result_last, today,my_domain):
    wb = Workbook()
    wb_all = Workbook()
    # 创建sheet写表头
    for group in group_list:
        sheet_num = 0
        wb.create_sheet(u'{0}'.format(group), index=sheet_num)
        wb_all.create_sheet(u'{0}'.format(group), index=sheet_num)
        row_first = ['日期', '首页', '总词数']
        row_first2 = ['日期', '域名','首页', '总词数']
        # 写表头
        wb[group].append(row_first)
        wb_all[group].append(row_first2)
        sheet_num += 1
    # 写内容
    for domain, dict_value in result_last.items():
        if domain == my_domain:
            for group, dict_index_all in dict_value.items():
                # 写数据
                row_value = [today]
                for key,value in dict_index_all.items():
                    row_value.append(value)
                wb[u'{0}'.format(group)].append(row_value)

        for group, dict_index_all in dict_value.items():
            # 写数据
            row_value = [today,domain]
            for key, value in dict_index_all.items():
                row_value.append(value)
            wb_all[u'{0}'.format(group)].append(row_value)
    wb.save('{0}bdmo1_index.xlsx'.format(today))
    wb_all.save('{0}bdmo1_index_domains.xlsx'.format(today))


def get_driver():
    c_service = Service(r'D:\install\pyhon36\chromedriver.exe')
    c_service.command_line_args()
    c_service.start()
    option = Options()
    option.binary_location = "C:/Program Files (x86)/Google/Chrome/Application/chrome.exe"  # 安装的位置
    # option.add_argument('disable-infobars')
    option.add_argument("--no-sandbox")
    option.add_argument("--disable-dev-shm-usage")
    option.add_argument("--disable-gpu")
    option.add_argument("--disable-features=NetworkService")
    # option.add_argument("--window-size=1920x1080")
    option.add_argument("--disable-features=VizDisplayCompositor")
    option.add_argument('headless')
    option.add_argument('log-level=3') #屏蔽日志
    No_Image_loading = {"profile.managed_default_content_settings.images": 2}
    option.add_experimental_option("prefs", No_Image_loading)
    driver = webdriver.Chrome(options=option, chrome_options=option)
    # 屏蔽特征
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": """
    Object.defineProperty(navigator, 'webdriver', {
      get: () => undefined
    })
  """
    })
    return driver,c_service


class bdmoIndexMonitor(threading.Thread):

    def __init__(self):
        threading.Thread.__init__(self)

    @staticmethod
    def read_excel(filepath):
        q = queue.Queue()
        group_list = []
        kwd_dict = {}
        wb_kwd = load_workbook(filepath)
        for sheet_obj in wb_kwd:
            sheet_name = sheet_obj.title
            group_list.append(sheet_name)
            kwd_dict[sheet_name]= []
            col_a = sheet_obj['A']
            for cell in col_a:
                kwd = (cell.value)
                # 加个判断吧
                if kwd:
                    q.put([sheet_name,kwd])
        return q, group_list

    # 初始化结果字典
    @staticmethod
    def result_init(group_list):
        result = {}
        for domain in domains:
            result[domain] = {}
            for group in group_list:
                result[domain][group] = {'首页':0,'总词数':0}
        print("结果字典init...")
        return result


    # 获取某词的serp源码上包含排名url的div块
    def get_divs(self, html ,url):
        div_list = []
        doc = pq(html)
        title = doc('title').text()
        if '- 百度' in title and 'https://m.baidu.com/s' in url:
            try:
                div_list = doc('.c-result').items()
            except Exception as e:
                print('提取div块失败', e)
            else:
                pass
        else:
            print('源码异常---------------------',title)
            time.sleep(120)
        return div_list

    # 提取排名的真实url
    def get_real_urls(self, div_list):
        real_urls_rank = []
        if div_list:
            for div in div_list:
                try:
                    data_log = div.attr('data-log')
                    data_log = json.loads(data_log.replace("'", '"')) # json字符串双引号
                    srcid = data_log['ensrcid'] if 'ensrcid' in data_log  else 'ensrcid' # 样式特征
                    rank_url = data_log['mu']  if 'mu' in data_log else None # mu可能为空或者不存在
                    rank = data_log['order']
                except Exception as e:
                    print('提取rank_url error',e)
                else:
                    if rank_url:
                        real_urls_rank.append((rank_url,rank,srcid))
                    # 如果mu为空或者不存在
                    else:
                        # 提取资讯聚合,图片聚合
                        article = div('.c-result-content article')
                        link = article.attr('rl-link-href')
                        # 提取热议聚合
                        if not link:
                            a = div('.c-result-content article header a')
                            data_log_ugc = a.attr('data-log')
                            data_log_ugc = json.loads(data_log_ugc.replace("'", '"')) # json字符串双引号
                            link = data_log_ugc['mu']  if 'mu' in data_log_ugc else None # mu可能为空或者不存在
                            link = 'https://m.baidu.com{0}'.format(link) if link != None else None
                            # 一般为卡片样式,链接太多,不提取了
                            if not link:
                                pass
                        real_urls_rank.append((link,rank,srcid))
        return real_urls_rank

    # 提取某url的域名部分
    def get_domain(self,real_url):
        domain = None
        try:
           res = urlparse(real_url)
        except Exception as e:
           print(e,'real_url:error')
        else:
           domain = res.netloc
        return domain

    # 获取某词serp源码首页排名所有域名
    def get_domains(self,real_url_list):
            domain_list = [self.get_domain(real_url) for real_url in real_url_list]
            # 一个词某域名多个url有排名,算一次
            domain_set = set(domain_list)
            domain_set = domain_set.remove(None) if None in domain_set else domain_set
            domain_str = ','.join(domain_set)
            return domain_str

    # 线程函数
    def run(self):
        global driver,c_service
        while 1:
            group_kwd = q.get()
            group,kwd = group_kwd
            print(group,kwd)
            try:
                driver.get('https://m.baidu.com/')
                input = WebDriverWait(driver, 30).until(
                    EC.visibility_of_element_located((By.ID, "index-kw"))
                )

                input_click_js = 'document.getElementById("index-kw").click()'
                driver.execute_script(input_click_js) # 点击输入框

                input_js = 'document.getElementById("index-kw").value="{0}"'.format(kwd)
                driver.execute_script(input_js) # 输入搜索词
                
                baidu = WebDriverWait(driver, 20).until(
                    EC.visibility_of_element_located((By.ID, "index-bn"))
                )
                click_js = 'document.getElementById("index-bn").click()'
                driver.execute_script(click_js) # 点击搜索

                # 等待首页元素加载完毕
                bottom = WebDriverWait(driver, 20).until(
                    EC.visibility_of_element_located((By.ID, "copyright"))
                )
                # 页面下拉
                driver.execute_script(js_xiala)
                html = driver.page_source
                now_url = driver.current_url
                divs_res = self.get_divs(html,now_url)
                # 源码ok再写入
                if divs_res:
                    real_urls_rank = self.get_real_urls(divs_res)
                    real_urls = []
                    for my_url,my_order,my_attr in real_urls_rank:
                        real_urls.append(my_url)
                        f_all.write('{0}\t{1}\t{2}\t{3}\t{4}\n'.format(kwd,my_url,my_order,my_attr,group))
                    f_all.flush()
                    domain_str = self.get_domains(real_urls)
                    # 目标站点是否出现
                    for domain in domains:
                        if domain not in domain_str:
                              f.write('{0}\t{1}\t{2}\t{3}\t{4}\n'.format(kwd, '无', '无', group,domain))
                        else:
                            for my_url,my_order,my_attr in real_urls_rank:
                                if domain in my_url:
                                    f.write('{0}\t{1}\t{2}\t{3}\t{4}\t{5}\n'.format(kwd,my_url,my_order,group,domain,my_attr))
                                    print(my_url, my_order)
                                    break # 取第一个排名url
                f.flush()
            except Exception as e:
                print(e)
                driver.quit()
                c_service.stop()
                driver,c_service = get_driver()

            finally:
                del kwd
                gc.collect()
                q.task_done()
                

if __name__ == "__main__":
    start = time.time()
    local_time = time.localtime()
    today = time.strftime('%Y%m%d',local_time)
    domains = ['5i5j.com','lianjia.com','anjuke.com','fang.com'] # 目标域名
    my_domain = '5i5j.com' # 自己域名
    js_xiala = 'window.scrollBy(0,{0} * {1})'.format('document.body.scrollHeight',random.random())
    driver,c_service = get_driver()
    
    
    q,group_list = bdmoIndexMonitor.read_excel('2020kwd_url_core_city_unique.xlsx')  # 关键词队列及分类
    result = bdmoIndexMonitor.result_init(group_list)  # 初始化结果
    all_num = q.qsize() # 总词数
    f = open('{0}bdmo1_index_info.txt'.format(today),'w',encoding="utf-8")
    f_all = open('{0}bdmo1_index_all.txt'.format(today),'w',encoding="utf-8")
    file_path = f.name
    # 设置线程数
    for i in list(range(1)):
        t = bdmoIndexMonitor()
        t.setDaemon(True)
        t.start()
    q.join()
    f.close()
    f_all.close()
    # 根据bdmo1_index_info.txt计算结果
    result_last = get_result(file_path,result)
    # 写入txt文件
    write_domains_txt(result_last)
    # 写入excel
    write_myexcel(group_list,result_last,today,my_domain)
     # 统计查询成功的词数
    with open(file_path,'r',encoding='utf-8') as fp:
         success = int(sum(1 for x in fp)/len(domains))
    end = time.time()
    print('关键词共{0}个,查询成功{1}耗时{2}min'.format(all_num,success, (end - start) / 60))

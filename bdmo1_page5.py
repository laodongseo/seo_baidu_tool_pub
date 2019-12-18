# ‐*‐ coding: utf‐8 ‐*‐
"""
分关键词种类批量查询网站首页/前二/三/四/五/五页覆盖率
bdmo1_page5_info.txt记录每个kwd在第几页有排名
bdmo1_page5_rankurl.txt记录每个kwd排名的url,当前页面有多个就记多个
bdmo1_page5.xlsx和bdmo1_page5.txt是统计结果
bdmo1_page5_info.txt的行数为查询成功词数
kwd_core_city.xlsx中sheet名代表关键词种类,每个sheet第一列放关键词
cookie必须是登录baidu账号后的cookie否则很容易被反爬
默认线程数2

"""

import requests
from pyquery import PyQuery as pq
import threading
import queue
import time
from urllib.parse import urlparse
from openpyxl import load_workbook
from openpyxl import Workbook
import time
import gc
import json


# 计算最终结果
def get_result(file_path, result, result_last):
    # 读取文件 计算首页到五页每页词数
    for line in open(file_path, 'r', encoding='utf-8'):
        line = line.strip().split('\t')
        page = line[1]
        group = line[2]
        result[group][page] += 1

    # 计算首页、前二页、前三页、前四页、前五页词数
    for group, page_num in result.items():
        print(group, page_num)
        for page, num in page_num.items():
            if page == '五页后':
                result_last[group]['五页后'] = num
            if page in ['首页', '二页', '三页', '四页', '五页']:
                result_last[group]['前五页'] += num
            if page in ['首页', '二页', '三页', '四页']:
                result_last[group]['前四页'] += num
            if page in ['首页', '二页', '三页']:
                result_last[group]['前三页'] += num
            if page in ['首页', '二页']:
                result_last[group]['前二页'] += num
            if page == '首页':
                result_last[group]['首页'] += num
    return result_last


# 写入excel文件
def write_excel(group_list, result_last, today):
    wb = Workbook()
    # 创建sheet
    for group in group_list:
        sheet_num = 0
        wb.create_sheet(u'{0}'.format(group), index=sheet_num)
        sheet_num += 1
    # 写内容
    row_first = ['日期', '首页', '前二页', '前三页', '前四页', '前五页', '五页后']
    for group, data_dict in result_last.items():
        # 写第一行表头
        wb[group].append(row_first)
        # 写数据
        row_value = [today]
        for page, value in data_dict.items():
            row_value.append(value)
        wb[u'{0}'.format(group)].append(row_value)
    wb.save('{0}bdmo1_page5.xlsx'.format(today))


class bdmoCoverPage5(threading.Thread):

    def __init__(self):
        threading.Thread.__init__(self)

    @staticmethod
    def read_excel(filepath):
        q = queue.Queue()
        group_list = []
        kwd_dict = {}
        wb_kwd = load_workbook(filepath)
        for sheet_obj in wb_kwd:
            sheet_name = sheet_obj.title
            group_list.append(sheet_name)
            kwd_dict[sheet_name]= []
            col_a = sheet_obj['A']
            for cell in col_a:
                kwd = (cell.value)
                # 加个判断吧
                if kwd:
                    q.put([sheet_name,kwd])
        return q, group_list

    # 初始化结果字典
    @staticmethod
    def result_init(group_list):
        pages = ['首页','二页','三页','四页','五页','五页后']
        result = {} #记录首页、第二页..第五页各多少词
        for group in group_list:
            result[group] = {}
        for group in group_list:
            for page in pages:
                result[group][page] = 0
        print("结果字典init...")
        return result

    # 初始化最后结果字典
    @staticmethod
    def result_last_init(group_list):
        result_last = {}  # 记录首页、第二页..第五页各多少词
        pages = ['首页','前二页','前三页','前四页','前五页','五页后']
        for group in group_list:
            result_last[group] = {}
        for group in group_list:
            for page in pages:
                result_last[group][page] = 0
        print("最终结果字典init...")
        return result_last

    # 获取某词serp源码
    def get_html(self,url,retry=2):
        try:
            r = requests.get(url=url,headers=user_agent,timeout=5)
        except Exception as e:
            print('获取源码失败',e)
            time.sleep(6)
            if retry > 0:
                self.get_html(url,retry-1)
        else:
            html = r.content.decode('utf-8',errors='ignore')  # 用r.text有时候识别错误
            url = r.url  # 反爬会重定向,取定向后的地址
            return html,url

    # 获取某词的serp源码上包含排名url的div块
    def get_data_logs(self, html ,url):
        data_logs = []
        doc = pq(html)
        title = doc('title').text()
        if '- 百度' in title and 'https://m.baidu.com/from=1001703y' in url:
            try:
                div_list = doc('.c-result').items()
            except Exception as e:
                print('提取div块失败', e)
            else:
                for div in div_list:
                    data_log = div.attr('data-log')
                    data_logs.append(data_log) if data_log is not None else data_logs
        else:
            print(title,'源码异常,可能反爬')
            time.sleep(100)
        return data_logs

    # 提取排名的真实url
    def get_real_urls(self, data_logs):
        real_urls = []
        for data_log in data_logs:
            # json字符串要以双引号表示
            data_log = json.loads(data_log.replace("'", '"'))
            url = data_log['mu']
            real_urls.append(url)
        return real_urls

    # 提取某url的域名部分
    def get_domain(self,real_url):
        domain = None
        try:
           res = urlparse(real_url)
        except Exception as e:
           print (e,real_url)
        else:
           domain = res.netloc
        return domain

    # 获取某词serp源码首页排名所有域名
    def get_domains(self,real_url_list):
            domain_list = [self.get_domain(real_url) for real_url in real_url_list]
            # 一个词某域名多个url有排名,计算一次
            domain_set = set(domain_list)
            domain_set = domain_set.remove(None) if None in domain_set else domain_set
            domain_str = ','.join(domain_set)
            return domain_str

    # 线程函数
    def run(self):
        while 1:
            group_kwd = q.get()
            group,kwd = group_kwd
            print(group,kwd)
            try:
                for page in page_dict.keys():
                    if page == '':
                        url = "https://m.baidu.com/from=1001703y/ssid=5ae0977cdc1ac126558a25f695277282.3.1576654032.1.6RmhZekZekV5/s?word={0}&ie=utf-8".format(kwd)
                    else:
                        url = "https://m.baidu.com/from=1001703y/ssid=5ae0977cdc1ac126558a25f695277282.3.1576654032.1.6RmhZekZekV5/s?pn={0}&usm=2&word={1}&ie=utf-8".format(kwd,page)
                    html,now_url = self.get_html(url)
                    data_logs = self.get_data_logs(html,now_url)
                    # 源码ok再写入
                    if data_logs:
                        real_url_list = self.get_real_urls(data_logs)
                        domain_str = self.get_domains(real_url_list)
                        if domain_str and domain in domain_str:
                            page = '首页' if page == '' else page_dict[page]
                            f.write('{0}\t{1}\t{2}\n'.format(kwd,page,group))
                            for real_url in real_url_list:
                                if domain in real_url:
                                    f_url.write('{0}\t{1}\t{2}\t{3}\n'.format(kwd,page,group,real_url))
                            break
                        if page == 40 and domain not in domain_str:
                            f.write('{0}\t{1}\t{2}\n'.format(kwd,'五页后',group))
                f.flush()
                f_url.flush()
            except Exception as e:
                print(e)
            finally:
                del kwd,group
                gc.collect()
                q.task_done()


if __name__ == "__main__":
    start = time.time()
    local_time = time.localtime()
    today = time.strftime('%Y%m%d',local_time)
    domain = '5i5j.com'
    user_agent = {
        'User-Agent': 'Mozilla/5.0 (Linux; Android 8.1.0; ALP-AL00 Build/HUAWEIALP-AL00; wv) AppleWebKit/537.36 (KHTML, like Gecko) Version/4.0 Chrome/63.0.3239.83 Mobile Safari/537.36 T7/10.13 baiduboxapp/10.13.0.11 (Baidu; P1 8.1.0)',
        'Referer': 'https://m.baidu.com/',
        'cookie':'wpr=0; BIDUPSID=95E739A8EE050812705C1FDE2584A61E; PSTM=1563865961; BAIDUID=95E739A8EE050812705C1FDE2584A61E:SL=0:NR=10:FG=1; BDUSS=NMRzZPVUFqR0JtbzJJc1ZDdkx2MGtiQUpvWVNUSjhnSUFmRFRmTnpDdmpGcXhkRVFBQUFBJCQAAAAAAAAAAAEAAADag5oxzI2IkAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAOOJhF3jiYRdOU; MSA_PBT=146; plus_lsv=f197ee21ffd230fd; plus_cv=1::m:49a3f4a6; MSA_ZOOM=1000; MSA_WH=394_670; BDORZ=FFFB88E999055A3F8A630C64834BD6D0; SE_LAUNCH=5%3A26276289_0%3A26276289; ysm=10277|10277; lsv=searchboxcss_591d86b-globalT_androidcss_1d3e059-wwwT_androidcss_42f8b70-globalBcss_14d95eb-wwwBcss_99b32fc; COOKIE_SESSION=25_0_0_1_0_t1_9_2_2_0_0_2_13_1576577784%7C2%230_0_0_0_0_0_0_0_1576577759%7C1; FC_MODEL=-1_0_5_0_4.71_1_0_0_1_0_4.71_-1_2_15_2_3_0_1576577980833_1576577784549%7C9%234.71_-1_-1_2_2_1576577980833_1576577784549%7C9; BDRCVFR[xoix5KwSHTc]=9xWipS8B-FspA7EnHc1QhPEUf; delPer=0; PSINO=3; H_PS_PSSID=; H_WISE_SIDS=136721_139419_139405_137831_114177_139148_120169_133995_138878_137979_137690_131247_132552_118880_118865_118839_118832_118793_138165_138882_136431_138845_138691_139283_139296_136863_138147_138114_139174_136195_131861_137105_139274_139400_139691_139430_133847_137734_138343_137467_138564_134256_131423_139396_138663_136537_110085_137441_127969_138302_137252_139507_139408_127417_138312_137187_136635_138425_139732_138943_135718_139221_139438_138753; rsv_i=a6aamxuScEJgZWD5o%2BojaSqSgR9aujTcvh1A4xMuiBzXRToQcCuTf8bzAhrmeTrgBRwomlxFgMtaszfud1CWLuo4e1agR%2F0; FEED_SIDS=345657_1218_14; BDSVRTM=412; BAIDULOC=11562630.22873027_149874.6866054242_16432_20001_1576654423165; H5LOC=1; Hm_lvt_12423ecbc0e2ca965d84259063d35238=1576231575,1576233002,1576577372,1576654423; Hm_lpvt_12423ecbc0e2ca965d84259063d35238=1576654423; ___rl__test__cookies=1576654423904; OUTFOX_SEARCH_USER_ID_NCOO=670906608.6866206; BDSVRBFE=Go; wise_tj_ub=ci%40103_21_-1_-1_-1_-1_-1%7Ciq%4014_1_13_232%7Ccb%40-1_-1_-1_-1_-1_-1_-1%7Cce%401%7Ctse%401; BDICON=10123156; __bsi=8491403584404635497_00_19_N_R_11_0303_c02f_Y'
        }
    q,group_list = bdmoCoverPage5.read_excel('kwd_core_city.xlsx')  # 关键词队列及分类
    result = bdmoCoverPage5.result_init(group_list)  # 结果字典-记录首页到5页各多少词
    result_last = bdmoCoverPage5.result_last_init(group_list)  # 结果字典-统计首页,前2页到前5页多少词
    all_num = q.qsize() # 总词数
    page_dict = {'':'首页',10:'二页',20:'三页',30:'四页',40:'五页'}  # 查询页码
    f = open('{0}bdmo1_page5_info.txt'.format(today),'w',encoding="utf-8")
    f_url = open('{0}bdmo1_page5_rankurl.txt'.format(today),'w',encoding="utf-8")
    file_path = f.name
    # 设置线程数
    for i in list(range(2)):
        t = bdmoCoverPage5()
        t.setDaemon(True)
        t.start()
    q.join()
    f.close()
    f_url.close()
    # 根据bdmo1_page5_info.txt数据计算结果
    result_last = get_result(file_path,result,result_last)
    # 写入txt文件
    with open('{0}bdmo1_page5.txt'.format(today), 'w', encoding="utf-8") as f:
        for group,data_dict in result_last.items():
            for page,value in data_dict.items():
                f.write(group + '\t' + page + '\t' + str(value) + '\n')
    end = time.time()
    # 写入excel
    write_excel(group_list,result_last,today)
    print('关键词共{0}个,成功个数看page5_info.txt,耗时{1}min'.format(all_num, (end - start) / 60))

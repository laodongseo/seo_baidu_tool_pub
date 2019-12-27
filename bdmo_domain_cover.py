# ‐*‐ coding: utf‐8 ‐*‐
"""
批量查询关键词采集移动端首页排名url计算各域名占比;
长期监控,观察行业流量的分发.
(如果某域名占比>100%,不是bug,说明有词出现多个url排名)
关键词文件kwd_xiaoqu_city.xlsx,每个sheet名代表关键词类别,可实现分关键词种类监控,sheet第一列放词
包含:相关网站.相关企业.智能小程序.其他人还在搜.热议聚合.资讯聚合.搜索智能聚合包含.视频(黄忠小区二手房)
不含:百度百科.百度手机助手下载
sigma.baidu.com:xx_相关网站|xx_相关企业
recommend_list.baidu.com:其他人还在搜
nourl.ubs.baidu.com:搜索智能聚合
bzclk.baidu.com:结构化的展示样式
/sf:xx-视频
res.txt是统计结果,url_serp.txt是所有的链接
"""

import requests
from pyquery import PyQuery as pq
import threading
import queue
import time
from urllib.parse import urlparse
import gc
import json
from openpyxl import load_workbook
from openpyxl import Workbook


# 提取某条url域名部分
def get_domain(real_url):

    # 通过mu提取url有些非自然排名url是空
    try:
       res = urlparse(real_url)  # real_url为空不会报错
    except Exception as e:
       print (e,real_url)
       domain = "xxx"
    else:
       domain = res.netloc
    return domain


# 获取某词serp源码首页排名真实url的域名部分
def get_domains(real_urls):
        domain_list = [get_domain(real_url) for real_url in real_urls]
        # 搜一个词 同一个域名多个url出现排名 只计算一次
        return domain_list

def save():
    res_format = result.items()
    #写入excel文件
    wb = Workbook()
    # 创建sheet
    for city in city_list:
        sheet_num = 0
        wb.create_sheet(u'{0}'.format(city),index=sheet_num)
        sheet_num += 1
    for city,data_dict in res_format:
        sort_dict = sorted(data_dict.items(), key=lambda s: s[1], reverse=True)
        for domain,num in sort_dict:
            row_value = [domain,num]
            wb[u'{0}'.format(city)].append(row_value)
    wb.save('{0}bdmo1_page5.xlsx'.format(today))

    # 写入txt
    res_format = sorted(result_all.items(), key=lambda s: s[1], reverse=True)
    with open('{0}res.txt'.format(today),'w',encoding='utf-8') as f:
        for domain,num in res_format:
            f.write(domain+'\t'+str(num)+'\n')

class bdmoCover(threading.Thread):

    def __init__(self):
        threading.Thread.__init__(self)

    # 读取文件 关键词进入队列
    @staticmethod
    def read_excel(filepath):
        q = queue.Queue()
        city_list = []
        wb_kwd = load_workbook(filepath)
        for sheet_obj in wb_kwd:
            sheet_name = sheet_obj.title
            city_list.append(sheet_name)
            col_a = sheet_obj['A']
            for cell in col_a:
                kwd = (cell.value)
                # 加个判断吧
                if kwd:
                    q.put([sheet_name,kwd])
        return q,city_list

    # 初始化结果字典
    @staticmethod
    def result_init(group_list):
        result = {}
        for group in group_list:
            result[group] = {}
        print("结果字典init...")
        return result

    # 获取某词serp源码
    def get_html(self,url,retry=2):
        try:
            r = requests.get(url=url,headers=user_agent,timeout=5)
        except Exception as e:
            print('获取源码失败',e)
            if retry > 0:
                self.get_html(url,retry-1)
        else:
            html = r.text
            return html,url

    # 获取某词的serp源码上包含排名url的div块
    def get_data_logs(self, html,url):
        data_logs = []
        url_other = []
        doc = pq(html)
        title = doc('title').text()
        if '- 百度' in title and 'https://m.baidu.com/ssid=da83cc8d88909a31' in url:
            try:
                div_list = doc('.c-result').items()
                # 如果mu为空,.c-result-content header a会有数据,否则没有
                a_list = doc('.c-result .c-result-content header a').items()
            except Exception as e:
                print('提取div块失败', e)
            else:
                for div in div_list:
                    data_log = div.attr('data-log')
                    data_logs.append(data_log) if data_log is not None else data_logs
                for a in a_list:
                    href = a.attr('data-sf-href')
                    url_other.append(href) if href is not None else href
        else:
            print('源码异常---------------------')
            time.sleep(120)
        return data_logs,url_other

    # 提取排名的真实url
    def get_real_urls(self, data_logs=[]):
        real_urls = []
        for data_log in data_logs:
            # json字符串要以双引号表示
            data_log = json.loads(data_log.replace("'", '"'))
            # 如果mu是空的话,real_urls里面会有空元素
            url = data_log['mu']
            real_urls.append(url)
        return real_urls


    # 线程函数
    def run(self):
        while 1:
            city,kwd = q.get()
            print(city,kwd,q.qsize())
            url = "https://m.baidu.com/ssid=da83cc8d88909a31/s?word={0}&ts=1818672&t_kt=0&ie=utf-8&rsv_iqid=1010771590&rsv_t=a17aiX76oe%252FA62s%252Fam2IHKuNujvGo8GqIRxwmX5UqwnhWUGUAlB2&sa=ib&rsv_pq=1010771590&rsv_sug4=2977&ss=100000000001&inputT=1157&sugid=6003555334805075793&tj=1".format(kwd)
            try:
                html,now_url = self.get_html(url)
                # f_url.write(html)
                data_logs,url_other = self.get_data_logs(html,now_url)
                real_urls = self.get_real_urls(data_logs)
                real_urls.extend(url_other)
                real_urls = [i for i in real_urls if i != '']
                for url_serp in real_urls:
                    f_url.write(kwd+'\t'+url_serp + '\t' + city + '\n')
                f_url.flush()
                del kwd
                gc.collect()
            except Exception as e:
                print(e)
            finally:
                q.task_done()



if __name__ == "__main__":
    start = time.time()
    local_time = time.localtime()
    today = time.strftime('%Y-%m-%d', local_time)
    f_url = open('url_serp{}.txt'.format(today),'w',encoding="utf-8")
    q,city_list = bdmoCover.read_excel('kwd_core_city.xlsx')
    result = bdmoCover.result_init(city_list)  # 结果字典区分城市统计
    result_all={} # 不区分城市统计
    user_agent = {
        'User-Agent': 'Mozilla/5.0 (Linux; Android 8.1.0; ALP-AL00 Build/HUAWEIALP-AL00; wv) AppleWebKit/537.36 (KHTML, like Gecko) Version/4.0 Chrome/63.0.3239.83 Mobile Safari/537.36 T7/10.13 baiduboxapp/10.13.0.11 (Baidu; P1 8.1.0)',
        'Referer': 'https://m.baidu.com/',
        'Cookie':'BDPASSGATE=IlPT2AEptyoA_yiU4V_43kIN8enzTri4H4PISkpT36ePdCyWmhTOAqVeEzChZm_YHi4hzp3afcJboCj-LzdXbaEddwYAhFpOakyM-M725aTvL0tdsbco_2z5V6g4trnG8vN7z4w1F3VEVFoKewPJpuo4ov4S9xFuku8O75LJrNO0YUOpDH8Pr7aTY767O-0APNu5-fqnhCpDMpihVPXkSS3Fg6chWp1L70aOatY6C3D5q6oY0RuiZMExGI8mFppi_x3nBQOLkKaoEV55qysc; max-age=43200; domain=.baidu.com; path=/'}


    #设置线程数
    for i in list(range(2)):
        t = bdmoCover()
        t.setDaemon(True)
        t.start()
    q.join()
    f_url.close()
    
    # 统计每个域名出现了多少次
    for i in open('{0}url_serp.txt'.format(today),'r',encoding='utf-8'):
        i = i.strip()
        line = i.split('\t')
        url = line[1]
        city = line[2]
        if url.startswith('http'):
            domain = get_domain(url)
            result[city][domain] = result[city][domain]+1 if domain in result[city] else 1
            result_all[domain] = result_all[domain]+1 if domain in result_all else 1
        if url.startswith('/sf'):
            result[city]['/sf'] = result[city]['/sf'] + 1 if '/sf' in result[city] else 1
            result_all['/sf'] = result_all['/sf'] + 1 if '/sf' in result_all else 1
    # 结果保存文件
    save()

    end = time.time()

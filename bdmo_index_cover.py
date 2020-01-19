# ‐*‐ coding: utf‐8 ‐*‐
"""
功能:分关键词类别批量查关键词采集移动端首页排名url计算各域名占比;
公式:某域名url个数/采集总url个数(一个词某域名可能有多个url排名计算在内)
说明:
   包含:相关网站.相关企业.智能小程序.其他人还在搜.热议聚合.资讯聚合.搜索智能聚合包含.视频(黄忠小区二手房)
   不含:百度手机助手下载的样式
   kwd_xiaoqu_city.xlsx:sheet名为关键词类别,sheet第一列放词
   http请求的cookie一定要用登录后的cookie
结果:
   res.txt:整体统计结果
   bdmo1_page5.xlsx 分关键词类别统计
   url_serp.txt:serp所有url
sigma.baidu.com:xx_相关网站|xx_相关企业
recommend_list.baidu.com:其他人还在搜
nourl.ubs.baidu.com:搜索智能聚合
bzclk.baidu.com:结构化的展示样式
/sf:某某-视频
"""

import requests
from pyquery import PyQuery as pq
import threading
import queue
import time
from urllib.parse import urlparse
import gc
import json
from openpyxl import load_workbook
from openpyxl import Workbook


# 提取某条url域名部分
def get_domain(real_url):

    # 通过mu提取url有些非自然排名url是空
    try:
       res = urlparse(real_url)  # real_url为空不会报错
    except Exception as e:
       print (e,real_url)
       domain = "xxx"
    else:
       domain = res.netloc
    return domain


# 采集某词首页所有排名真实url域名部分
def get_domains(real_urls):
        domain_list = [get_domain(real_url) for real_url in real_urls]
        # 搜一个词 同一个域名多个url出现排名 只计算一次
        return domain_list

def save():
    res_format = result.items()
    #写入excel文件
    wb = Workbook()
    # 创建sheet
    for city in city_list:
        sheet_num = 0
        wb.create_sheet(u'{0}'.format(city),index=sheet_num)
        sheet_num += 1
    for city,data_dict in res_format:
        sort_dict = sorted(data_dict.items(), key=lambda s: s[1], reverse=True)
        for domain,num in sort_dict:
            row_value = [domain,num]
            wb[u'{0}'.format(city)].append(row_value)
    wb.save('{0}bdmo1_page5.xlsx'.format(today))

    # 写入txt
    res_format = sorted(result_all.items(), key=lambda s: s[1], reverse=True)
    with open('{0}res.txt'.format(today),'w',encoding='utf-8') as f:
        for domain,num in res_format:
            f.write(domain+'\t'+str(num)+'\n')

class bdmoCover(threading.Thread):

    def __init__(self):
        threading.Thread.__init__(self)

    # 读取文件 关键词进入队列
    @staticmethod
    def read_excel(filepath):
        q = queue.Queue()
        city_list = []
        wb_kwd = load_workbook(filepath)
        for sheet_obj in wb_kwd:
            sheet_name = sheet_obj.title
            city_list.append(sheet_name)
            col_a = sheet_obj['A']
            for cell in col_a:
                kwd = (cell.value)
                # 加个判断,防止一些不可见字符
                if kwd:
                    kwd_z = kwd + '租房'
                    kwd_er = kwd + '二手房'
                    q.put([sheet_name,kwd_z])
                    q.put([sheet_name,kwd_er])
        return q,city_list

    # 初始化结果字典
    @staticmethod
    def result_init(group_list):
        result = {}
        for group in group_list:
            result[group] = {}
        print("结果字典init...")
        return result

    # 获取某词serp源码
    def get_html(self,url,retry=2):
        try:
            r = requests.get(url=url,headers=my_header,timeout=5)
        except Exception as e:
            print('获取源码失败',e)
            if retry > 0:
                self.get_html(url,retry-1)
        else:
            html = r.text
            return html,url

    # 获取某词的serp源码上包含排名url的div块
    def get_data_logs(self, html,url):
        data_logs = []
        url_other = []
        doc = pq(html)
        title = doc('title').text()
        if '- 百度' in title and 'https://m.baidu.com/ssid=da83cc8d88909a31' in url:
            try:
                div_list = doc('.c-result').items()
                # 如果mu为空,.c-result-content header a会有数据,这类数据样式特别,比如资讯聚合
                a_list = doc('.c-result .c-result-content header a').items()
            except Exception as e:
                print('提取div块失败', e)
            else:
                for div in div_list:
                    data_log = div.attr('data-log')
                    data_logs.append(data_log) if data_log is not None else data_logs
                for a in a_list:
                    href = a.attr('data-sf-href')
                    url_other.append(href) if href is not None else href
        else:
            print('源码异常---------------------')
            time.sleep(120)
        return data_logs,url_other

    # 提取排名的真实url
    def get_real_urls(self, data_logs=[]):
        real_urls = []
        for data_log in data_logs:
            # json字符串要以双引号表示
            data_log = json.loads(data_log.replace("'", '"'))
            # 如果mu是空的话,real_urls里面会有空元素
            url = data_log['mu']
            real_urls.append(url)
        return real_urls


    # 线程函数
    def run(self):
        while 1:
            city,kwd = q.get()
            print(city,kwd,q.qsize())
            url = "https://m.baidu.com/ssid=da83cc8d88909a31/s?word={0}&ts=1818672&t_kt=0&ie=utf-8&rsv_iqid=1010771590&rsv_t=a17aiX76oe%252FA62s%252Fam2IHKuNujvGo8GqIRxwmX5UqwnhWUGUAlB2&sa=ib&rsv_pq=1010771590&rsv_sug4=2977&ss=100000000001&inputT=1157&sugid=6003555334805075793&tj=1".format(kwd)
            try:
                html,now_url = self.get_html(url)
                # f_url.write(html)
                data_logs,url_other = self.get_data_logs(html,now_url)
                real_urls = self.get_real_urls(data_logs)
                real_urls.extend(url_other)
                real_urls = [i for i in real_urls if i != '']
                for url_serp in real_urls:
                    f_url.write(kwd+'\t'+url_serp + '\t' + city + '\n')
                f_url.flush()
                del kwd
                gc.collect()
            except Exception as e:
                print(e)
            finally:
                q.task_done()



if __name__ == "__main__":
    start = time.time()
    local_time = time.localtime()
    today = time.strftime('%Y-%m-%d', local_time)
    f_url = open('{}url_serp.txt'.format(today),'w',encoding="utf-8")
    q,city_list = bdmoCover.read_excel('kwd_core_city.xlsx')
    result = bdmoCover.result_init(city_list)  # 初始化结果字典
    result_all={} # 不区分城市统计
    # head设置
    my_header = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
        'Connection': 'keep-alive',
        'Cookie': 'BIDUPSID=95E739A8EE050812705C1FDE2584A61E; PSTM=1563865961; BAIDUID=95E739A8EE050812705C1FDE2584A61E:SL=0:NR=10:FG=1; BDUSS=NMRzZPVUFqR0JtbzJJc1ZDdkx2MGtiQUpvWVNUSjhnSUFmRFRmTnpDdmpGcXhkRVFBQUFBJCQAAAAAAAAAAAEAAADag5oxzI2IkAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAOOJhF3jiYRdOU; MSA_PBT=146; plus_lsv=f197ee21ffd230fd; plus_cv=1::m:49a3f4a6; MSA_ZOOM=1056; MSA_WH=414_736; COOKIE_SESSION=2855_6_3_6_4_m1_7_8_0_0_1_6_48_1578649761%7C9%230_0_0_0_0_0_0_0_1578636926%7C1; FC_MODEL=-1_6_3_0_7.28_0_3_0_0_0_283.07_-1_7_7_6_11_0_1578650188874_1578649761311%7C9%237.28_-1_-1_7_6_1578650188874_1578649761311%7C9; H_PS_PSSID=; BDORZ=FFFB88E999055A3F8A630C64834BD6D0; SIGNIN_UC=70a2711cf1d3d9b1a82d2f87d633bd8a03288920344; delPer=0; PSINO=3; H_WISE_SIDS=141144_139203_139419_139403_137831_114177_139396_135846_139148_120169_139864_140833_133995_138878_137979_140173_131247_132552_140227_118880_118865_118839_118832_118793_138165_107317_138883_140260_136431_139046_140592_138147_140120_139174_139624_140114_136196_131861_137105_140591_139694_138586_133847_140792_137734_140545_134256_131423_140823_138663_136537_141103_110085_140325_127969_140622_140595_140864_139802_137252_139408_127417_138312_138425_139733_139912_140685_139926_140596_138754_140964; FEED_SIDS=345657_0113_8; Hm_lvt_12423ecbc0e2ca965d84259063d35238=1578632578,1578640885,1578646881,1578880697; SE_LAUNCH=5%3A26314678_0%3A26314678; rsv_i=0145AUxG3XFh5hste9dXED%2FtBt7QdwocVgy8MNIDFD%2FqxXut9LoO0cJ84FJvz3WoYkVml5f%2BBfxYrNkd2WvGrJsFIbnDBh4; BDSVRTM=384; Hm_lpvt_12423ecbc0e2ca965d84259063d35238=1578880700; ___rl__test__cookies=1578880924415; OUTFOX_SEARCH_USER_ID_NCOO=226708085.36203498; BDSVRBFE=Go; wise_tj_ub=ci%4089_44_89_44_90_137_54%7Ciq%403_4_3_511%7Ccb%40-1_-1_-1_-1_-1_-1_-1%7Cce%401%7Ctse%401; __bsi=8932389214787994849_00_8_R_R_8_0303_c02f_Y; BDICON=10123156',
        'Host': 'm.baidu.com',
        'Referer': 'https://m.baidu.com/',
        'Sec-Fetch-Mode': 'navigate',
        'Sec-Fetch-Site': 'same-origin',
        'Sec-Fetch-User': '?1',
        'Upgrade-Insecure-Requests': '1',
        'User-Agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 11_0 like Mac OS X) AppleWebKit/604.1.38 (KHTML, like Gecko) Version/11.0 Mobile/15A372 Safari/604.1'}


    #设置线程数
    for i in list(range(2)):
        t = bdmoCover()
        t.setDaemon(True)
        t.start()
    q.join()
    f_url.close()
    # 统计每个域名出现了多少次
    for i in open('{0}url_serp.txt'.format(today),'r',encoding='utf-8'):
        i = i.strip()
        line = i.split('\t')
        url = line[1]
        city = line[2]
        if url.startswith('http'):
            domain = get_domain(url)
            result[city][domain] = result[city][domain]+1 if domain in result[city] else 1
            result_all[domain] = result_all[domain]+1 if domain in result_all else 1
        if url.startswith('/sf'):# 独立提取出来
            result[city]['/sf'] = result[city]['/sf'] + 1 if '/sf' in result[city] else 1
            result_all['/sf'] = result_all['/sf'] + 1 if '/sf' in result_all else 1
    # 结果保存文件
    save()

    end = time.time()

# ‐*‐ coding: utf‐8 ‐*‐
"""
功能:
   1)指定几个域名,分关键词种类监控首页词数
   2)抓取serp所有url,提取域名并统计各域名首页覆盖率
   3)通过tpl属性记录serp排名url的特征
提示:
  1)含自然排名和百度开放平台的排名
  2)百度开放平台的样式mu属性值为排名url,mu不存在提取article里的url
  3)2020xiaoqu_kwd_city_new.xlsx:sheet名为关键词种类,sheet第一列放关键词
结果:
    bdpc1_index_info.txt:各监控站点词的排名及url,如有2个url排名,只取第一个
    bdpc1_index_all.txt:serp所有url及样式特征,依此统计各域名首页覆盖率-单写脚本统计
    bdpc1_index.xlsx:自己站每类词首页词数
    bdpc1_index_domains.xlsx:各监控站点每类词的首页词数
    bdpc1_index_domains.txt:各监控站点每类词的首页词数
header头信息复制浏览器全部的请求头,Accept-Encoding留deflate
serp上请求头多测几个找到和手动搜索结果最一致的

"""

import requests
from pyquery import PyQuery as pq
import threading
import queue
import time
from urllib.parse import urlparse
from openpyxl import load_workbook
from openpyxl import Workbook
import time
import gc
import random


# 计算最终结果
def get_result(file_path, result):
    for line in open(file_path, 'r', encoding='utf-8'):
        line = line.strip().split('\t')
        rank = line[2]
        group = line[3]
        domain = line[4]
        if rank != '无':
            result[domain][group]['首页'] += 1
        result[domain][group]['总词数'] += 1
    return result

# 写txt,所有监控域名的结果
def write_domains_txt(result_last):
    with open('{0}bdpc1_index_domains.txt'.format(today), 'w', encoding="utf-8") as f_res:
        f_res.write('{0}\t{1}\t{2}\t{3}\t{4}\n'.format('日期','域名','词类','首页词数','查询词数'))
        for now_domain,dict_value in result_last.items():
            for group, dict_index_all in dict_value.items():
                f_res.write('{0}\t{1}\t{2}\t'.format(today,now_domain,group))
                for key, value in dict_index_all.items():
                    f_res.write(str(value) + '\t')
                f_res.write('\n')


# 写excel
def write_myexcel(group_list, result_last, today,my_domain):
    wb = Workbook()
    wb_all = Workbook()
    # 创建sheet写表头
    for group in group_list:
        sheet_num = 0
        wb.create_sheet(u'{0}'.format(group), index=sheet_num)
        wb_all.create_sheet(u'{0}'.format(group), index=sheet_num)
        row_first = ['日期', '首页', '总词数']
        row_first2 = ['日期', '域名','首页', '总词数']
        # 写表头
        wb[group].append(row_first)
        wb_all[group].append(row_first2)
        sheet_num += 1
    # 写内容
    for domain, dict_value in result_last.items():
        if domain == my_domain:
            for group, dict_index_all in dict_value.items():
                # 写数据
                row_value = [today]
                for key,value in dict_index_all.items():
                    row_value.append(value)
                wb[u'{0}'.format(group)].append(row_value)

        for group, dict_index_all in dict_value.items():
            # 写数据
            row_value = [today,domain]
            for key, value in dict_index_all.items():
                row_value.append(value)
            wb_all[u'{0}'.format(group)].append(row_value)
    wb.save('{0}bdpc1_index.xlsx'.format(today))
    wb_all.save('{0}bdpc1_index_domains.xlsx'.format(today))

# 发js包-不用
def request_js(url,my_header,retry=1):
    try:
        r = requests.get(url=url,headers=my_header,timeout=2)
    except Exception as e:
        print('获取源码失败',e)
        time.sleep(6)
        if retry > 0:
            request_js(url,retry-1)
    else:
        pass

# header
def get_header():
    my_header = {
       'Accept':'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
'Accept-Encoding':'deflate',
'Accept-Language':'zh-CN,zh;q=0.9,en;q=0.8',
'Cache-Control':'no-cache',
'Connection':'keep-alive',
'Cookie':'BIDUPSID=F85975B8C2BE20C8CF7C2910258B6AAB; PSTM=1587465347; BAIDUID=F85975B8C2BE20C81F0639FB4F15217F:FG=1; BD_HOME=1; BD_UPN=12314353; delPer=0; BD_CK_SAM=1; PSINO=1; H_PS_PSSID=30974_1443_31125_21109_31069_31341_30823_26350_31163_31196; BDORZ=B490B5EBF6F3CD402E515D22BCDA1598; H_PS_645EC=d43eS%2BEeOoE00584LmGORoxGJ3euKFsJ1YPm8TwUtTgzKksN1rqUhLoCOK4; BDSVRTM=108; COOKIE_SESSION=62_0_2_1_0_1_0_0_1_1_2_0_0_0_0_0_0_0_1587465449%7C3%230_0_1587465449%7C1',
'Host':'www.baidu.com',
'Pragma':'no-cache',
'Sec-Fetch-Dest':'document',
'Sec-Fetch-Mode':'navigate',
'Sec-Fetch-Site':'same-origin',
'Sec-Fetch-User':'?1',
'Upgrade-Insecure-Requests':'1',
'User-Agent':'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.163 Safari/537.36',}
    return my_header



class bdpcIndexMonitor(threading.Thread):

    def __init__(self):
        threading.Thread.__init__(self)

    @staticmethod
    def read_excel(filepath):
        q = queue.Queue()
        group_list = []
        kwd_dict = {}
        wb_kwd = load_workbook(filepath)
        for sheet_obj in wb_kwd:
            sheet_name = sheet_obj.title
            group_list.append(sheet_name)
            kwd_dict[sheet_name]= []
            col_a = sheet_obj['B']
            for cell in col_a:
                kwd = (cell.value)
                # 加个判断吧
                if kwd:
                    q.put([sheet_name,kwd])
        return q, group_list

    # 初始化结果字典
    @staticmethod
    def result_init(group_list):
        result = {}
        for domain in domains:
            result[domain] = {}
            for group in group_list:
                result[domain][group] = {'首页':0,'总词数':0}
        print("结果字典init...")
        return result


    # 获取某词serp源码
    def get_html(self,url,my_header,retry=1):
        try:
            r = requests.get(url=url,headers=my_header,timeout=5)
        except Exception as e:
            print('获取源码失败',e)
            time.sleep(6)
            if retry > 0:
                self.get_html(url,my_header,retry-1)
        else:
            html = r.content.decode('utf-8',errors='ignore')  # 用r.text有时候识别错误
            url = r.url  # 反爬会重定向,取定向后的地址
            return html,url

    # 获取某词serp源码所有url
    def get_encrpt_urls(self,html,url):
        encrypt_url_list = []
        real_urls = []
        doc = pq(html)
        title = doc('title').text()
        if '_百度搜索' in title and 'https://www.baidu.com/s?ie=utf-8' in url:
            div_list = doc('.result').items() # 自然排名
            div_op_list = doc('.result-op').items() # 非自然排名
            for div in div_list:
                rank = div.attr('id')
                tpl = div.attr('tpl') if div.attr('tpl') else 'xxx'
                if rank:
                    try:
                        a = div('h3 a')
                    except Exception as e:
                        print('未提取自然排名加密链接')
                    else:
                        encrypt_url = a.attr('href')
                        encrypt_url_list.append((encrypt_url,rank,tpl))
            for div in div_op_list:
                rank_op = div.attr('id')
                tpl = div.attr('tpl')
                tpl = div.attr('tpl') if div.attr('tpl') else 'xxx'
                if rank_op:
                    link = div.attr('mu') # 真实url,有些op样式没有mu属性
                    # print(link,rank_op)
                    if link: 
                        real_urls.append((link,rank_op,tpl))
                    else:
                        encrypt_url = div('article a').attr('href')
                        encrypt_url_list.append((encrypt_url,rank_op,tpl))

        else:
            print('源码异常,可能反爬')
            print(html)
            time.sleep(60)
                    
        return encrypt_url_list,real_urls

    # 解密某条加密url
    def decrypt_url(self,encrypt_url,my_header,retry=1):
        real_url = None # 默认None
        try:
            encrypt_url = encrypt_url.replace('http://','https://')
            r = requests.head(encrypt_url,headers=my_header)
        except Exception as e:
            print(encrypt_url,'解密失败',e)
            time.sleep(6)
            if retry > 0:
                self.decrypt_url(encrypt_url,my_header,retry-1)
        else:
            real_url = r.headers['Location'] if 'Location' in r.headers else None
        return real_url

    # 获取某词serp源码首页排名真实url
    def get_real_urls(self,encrypt_url_list):
        real_url_list = [self.decrypt_url(encrypt_url) for encrypt_url in encrypt_url_list]
        real_url_set = set(real_url_list)
        real_url_set = real_url_set.remove(None) if None in real_url_set else real_url_set
        real_url_list = list(real_url_set)
        return real_url_list

    # 提取某url的域名部分
    def get_domain(self,real_url):
        domain = None
        try:
           res = urlparse(real_url)
        except Exception as e:
           print (e,real_url)
        else:
           domain = res.netloc
        return domain

    # 获取某词serp源码首页排名所有域名
    def get_domains(self,real_url_list):
            domain_list = [self.get_domain(real_url) for real_url in real_url_list]
            # 一个词某域名多个url有排名,计算一次
            domain_set = set(domain_list)
            domain_set = domain_set.remove(None) if None in domain_set else domain_set
            domain_str = ','.join(domain_set)
            return domain_str

    # 线程函数
    def run(self):
        while 1:
            group_kwd = q.get()
            group,kwd = group_kwd
            print(group,kwd)
            try:
                url = "https://www.baidu.com/s?ie=utf-8&rsv_bp=1&wd={0}".format(kwd)
                my_header = get_header()
                html,now_url = self.get_html(url,my_header)
                encrypt_url_list_rank,real_urls_rank = self.get_encrpt_urls(html,now_url)
                # 源码ok再写入
                if encrypt_url_list_rank:
                    for my_serp_url,my_order,tpl in encrypt_url_list_rank:
                        my_real_url = self.decrypt_url(my_serp_url,my_header)
                        real_urls_rank.append((my_real_url,my_order,tpl))
                    real_urls = []
                    for my_real_url,my_order,tpl in real_urls_rank:
                        real_urls.append(my_real_url)
                        # print(tpl)
                        f_all.write('{0}\t{1}\t{2}\t{3}\t{4}\n'.format(kwd,my_real_url,my_order,tpl,group))
                    domain_str = self.get_domains(real_urls)
                    # 目标站点是否出现
                    for domain in domains:
                        if domain not in domain_str:
                            f.write('{0}\t{1}\t{2}\t{3}\t{4}\n'.format(kwd, '无', '无', group,domain))
                        else:
                            for my_url, my_order,tpl in real_urls_rank:
                                if domain in my_url:
                                    f.write('{0}\t{1}\t{2}\t{3}\t{4}\t{5}\n'.format(kwd, my_url, my_order, group,domain,tpl))
                                    # print(my_url, my_order)
                                    break # 取第一个排名url
                f.flush()
                f_all.flush()
            except Exception as e:
                print(e)
            finally:
                del kwd
                gc.collect()
                q.task_done()


if __name__ == "__main__":
    start = time.time()
    local_time = time.localtime()
    today = time.strftime('%Y%m%d',local_time)
    domains = ['5i5j.com','lianjia.com','anjuke.com','fang.com'] # 目标域名
    my_domain = '5i5j.com'
    q,group_list = bdpcIndexMonitor.read_excel('2020xiaoqu_kwd_city_new.xlsx')  # 关键词队列及分类
    result = bdpcIndexMonitor.result_init(group_list)  # 结果字典
    # print(result)
    all_num = q.qsize() # 总词数
    f = open('{0}bdpc1_index_info.txt'.format(today),'w',encoding="utf-8")
    f_all = open('{0}bdpc1_index_all.txt'.format(today),'w',encoding="utf-8")
    file_path = f.name
    # 设置线程数
    for i in list(range(1)):
        t = bdpcIndexMonitor()
        t.setDaemon(True)
        t.start()
    q.join()
    f.close()
    f_all.close()
    # 根据bdpc1_index_info.txt计算结果
    result_last = get_result(file_path,result)
    # 写入txt文件
    write_domains_txt(result_last)
    # 写入excel
    write_myexcel(group_list,result_last,today,my_domain)
    end = time.time()
    print('关键词共{0}个,耗时{1}min'.format(all_num, (end - start) / 60))

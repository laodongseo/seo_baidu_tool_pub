# ‐*‐ coding: utf‐8 ‐*‐
"""
评估网站和竞品的流量比
给所有搜索词设个平均搜索量,计算如下：
域名流量 = (排名第1词数 * 第1名点击率 + 排名第2词数 * 第2名点击率 +...+ 排名第10词数 * 第10名点击率) * 平均搜索量
domains.txt放要对比的域名,一行一个
bdpc_serp.txt每行以json的形式记录一个关键词首页自然排名url及排名值
kwd_core_city.xlsx放关键词,可以多个sheet,每个sheet代表一类词,每个sheet用第一列
bdpc_res.xlsx为结果，记录每个域名排名1—10的关键词个数
(一个词同个域名出现两个url排名会全部记录)
线程数默认2
"""

import requests
from pyquery import PyQuery as pq
import threading
import queue
import time
import json
from openpyxl import load_workbook
from openpyxl import Workbook
import random


# 初始化结果字典,每个域名1-10名排名个数全为0
def make_result(domains):
    result = {}
    for domain in domains:
        # 初始1—10名排名个数全为0个,不能写在循环外否则会产生深拷贝浅拷贝的问题
        init_value = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0}
        result[domain] = init_value
    return result


# 各域名1-10名排名个数写入到excel
def save_excel(result):
    wb = Workbook()
    ws = wb.active
    list_cell = ['domain']
    list_cell.extend([i for i in range(1, 11)])  # 第一行表头:域名及1-10数
    ws.append(list_cell)
    for domain, kv in result.items():
        row_value = [domain]
        ranks = list(kv.values())
        row_value.extend(ranks)  # 每行数据:域名及1-10排名数据
        ws.append(row_value)
    wb.save('{0}bdpc_res.xlsx'.format(today))


class BdpcLiuLiang(threading.Thread):

    def __init__(self):
        threading.Thread.__init__(self)

    # 读取excel文件获取关键词
    @staticmethod
    def read_excel(filepath):
        q = queue.Queue()
        wb_kwd = load_workbook(filepath)
        for sheet_obj in wb_kwd:
            col_a = sheet_obj['A']  # 默认取sheet第一列
            for cell in col_a:
                kwd = (cell.value)
                # 加个判断,防止一些不可见字符
                if kwd:
                    q.put(kwd)
        return q

    # 获取某词的serp源码
    def get_html(self, url, retry=2):
        try:
            user_agent = random.choice(my_header)
            r = requests.get(url=url, headers=user_agent, timeout=5)
        except Exception as e:
            print('获取源码失败', url, e)
            if retry > 0:
                self.get_html(url, retry - 1)
        else:
            html = r.content.decode('utf-8',errors='ignore')
            url = r.url
            return html,url

    # 获取某词的serp源码上自然排名的所有url
    def get_encrpt_urls(self, html,url):
        encrypt_url_list = []
        doc = pq(html)
        title = doc('title').text()
        if '_百度搜索' in title and 'https://www.baidu.com/s?tn=48020221' in url:
            div_list = doc('.result').items() # 自然排名/百度快照
            # div_op_list = doc('.result-op').items() # 非自然排名
            for div in div_list:
                rank = div.attr('id')
                if rank:
                    try:
                        a_list = div('.t a').items()
                    except Exception as e:
                        print('未提取自然排名加密链接')
                    else:
                        for a in a_list:
                            encrypt_url = a.attr('href')
                            if encrypt_url.find('http://www.baidu.com/link?url=') == 0:
                                encrypt_url_list.append((encrypt_url,rank))
        else:
            print('源码异常,可能反爬')
            print(html)
            time.sleep(6)
                    
        return encrypt_url_list

    # 解密某条加密url
    def decrypt_url(self, encrypt_url, retry=1):
        url = 'xxxxxxxxxxx'
        try:
            encrypt_url = encrypt_url.replace('http://', 'https://')
            r = requests.head(encrypt_url, headers=random.choice(my_header))
        except Exception as e:
            print(encrypt_url, '解密失败', e)
            time.sleep(5)
            if retry > 0:
                self.decrypt_url(encrypt_url, retry - 1)
        else:
            url = r.headers['Location']
            return url

    # 线程函数
    def run(self):
        while 1:
            kwd = q.get()
            url = "https://www.baidu.com/s?tn=48020221_28_hao_pg&ie=utf-8&wd={0}".format(kwd)
            try:
                threadLock.acquire() # 一行分多次写入,加锁避免数据错乱
                html,now_url = self.get_html(url)
                encrypt_url_list_rank = self.get_encrpt_urls(html,now_url)
                if encrypt_url_list_rank:
                    f.write('{{"{0}":['.format(kwd))  # 不是匹配符号的{} 需要双倍写
                    length = len(encrypt_url_list_rank)
                    num = 0
                    for encrypt_url_rank in encrypt_url_list_rank:
                        encrypt_url,rank= encrypt_url_rank[0],encrypt_url_rank[1]
                        url = self.decrypt_url(encrypt_url)
                        print(kwd,url,rank)
                        num += 1
                        json_str = str('{{"{0}":"{1}"}},'.format(url,rank))  if num < length else str('{{"{0}":"{1}"}}'.format(url,rank))
                        f.write(json_str)
                    f.write(']}\n')
                    f.flush()
            except Exception as e:
                print(e)
            finally:
                threadLock.release()
                q.task_done()


if __name__ == "__main__":
    start = time.time()
    today = time.strftime('%Y%m%d', time.localtime())
    # 目标监控域名
    target_domains = [i.strip() for i in open('domains.txt','r')]
    # 结果字典
    result = make_result(target_domains)
    threadLock = threading.Lock()
    my_header = [
        {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36',
         'Cookie': 'BIDUPSID=95E739A8EE050812705C1FDE2584A61E; PSTM=1563865961; BAIDUID=95E739A8EE050812705C1FDE2584A61E:SL=0:NR=10:FG=1; BDUSS=NMRzZPVUFqR0JtbzJJc1ZDdkx2MGtiQUpvWVNUSjhnSUFmRFRmTnpDdmpGcXhkRVFBQUFBJCQAAAAAAAAAAAEAAADag5oxzI2IkAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAOOJhF3jiYRdOU; NOJS=1; BD_UPN=12314353; MSA_WH=346_288; BAIDULOC=11562630.22873027_149874.6866054242_16432_20001_1574738954142; H_WISE_SIDS=137735_138198_137935_137831_114177_137657_135847_136436_137971_120169_138490_137381_133995_137979_132909_137690_131247_132552_137750_136680_118880_118865_118839_118832_118793_107313_136431_133352_137900_136863_138147_136194_131861_137105_133847_138476_138343_137467_138648_131423_138511_136537_138178_110085_137441_127969_137912_137829_138152_127417_136635_138425_135718_138302_137449_138239; COOKIE_SESSION=3663_0_8_1_3_6_0_3_7_2_0_5_19701_0_5_0_1574753671_0_1574753666%7C9%231036863_54_1573608641%7C9; cflag=13%3A3; pgv_pvi=7288111104; BDRCVFR[xoix5KwSHTc]=9xWipS8B-FspA7EnHc1QhPEUf; delPer=0; BD_CK_SAM=1; PSINO=3; H_PS_PSSID=; sug=0; sugstore=1; ORIGIN=0; bdime=20100; BDORZ=FFFB88E999055A3F8A630C64834BD6D0; BDSVRTM=0',
         'Referer': 'https://www.hao123.com/?tn=48020221_28_hao_pg'},
        ]
    q = BdpcLiuLiang.read_excel('kwd_core_city.xlsx')
    f = open('{0}bdpc_serp.txt'.format(today),'w',encoding='utf-8')  # 保存serp结果数据
    # 设置线程数
    for i in list(range(2)):
        t = BdpcLiuLiang()
        t.setDaemon(True)
        t.start()
    q.join()
    f.flush()
    f.close()

    # 读取文件,计算结果
    for line in open('{0}bdpc_serp.txt'.format('20191128'),'r',encoding='utf-8'):
        try:
            dic = json.loads(line.strip())  # 转为字典
            values = list(dic.values())[0]  # 获取值:各域名及排名值
        except Exception as e:
            print(e)
        else:
            # 循环每行中所有自然排名url及rank,判断属于哪个域名
            for kv_dict in values:
                url,rank = list(kv_dict.keys())[0],list(kv_dict.values())[0]
                for domain in target_domains: # 一个词同域名出现2个url排名,全算在内
                    print('当前排名url是{0},监控域名是{1}'.format(url,domain))
                    if domain in url:
                        result[domain][int(rank)] += 1
                        break
    save_excel(result)
    end = time.time()
    print('end...成功查询个数为txt文件行数,耗时{} min'.format((end-start)/60))

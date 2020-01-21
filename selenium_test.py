# ‐*‐ coding: utf‐8 ‐*‐
"""
分关键词种类批量查询首页/前二/三/四/五/五页词数
bdmo1_page5_info.txt记录每个kwd在第几页有排名
bdmo1_page5_rankurl.txt记录每个kwd排名的url,当前页面出现多个就记多个
bdmo1_page5.xlsx和bdmo1_page5.txt是统计结果
bdmo1_page5_info.txt的行数为查询成功词数
kwd_core_city.xlsx中sheet名代表关键词种类,每个sheet第一列放关键词
content = gzip.decompress(resp.text).decode('utf-8','ignore')
用的selenium驱动浏览器
"""

import requests
from pyquery import PyQuery as pq
import threading
import queue
import time
from urllib.parse import urlparse
from openpyxl import load_workbook
from openpyxl import Workbook
import time
import gc
import json
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options


# 计算最终结果
def get_result(file_path, result, result_last):
    # 读取文件 计算首页到五页每页词数
    for line in open(file_path, 'r', encoding='utf-8'):
        line = line.strip().split('\t')
        page = line[1]
        group = line[2]
        result[group][page] += 1

    # 计算首页、前二页、前三页、前四页、前五页词数
    for group, page_num in result.items():
        print(group, page_num)
        for page, num in page_num.items():
            if page == '五页后':
                result_last[group]['五页后'] = num
            if page in ['首页', '二页', '三页', '四页', '五页']:
                result_last[group]['前五页'] += num
            if page in ['首页', '二页', '三页', '四页']:
                result_last[group]['前四页'] += num
            if page in ['首页', '二页', '三页']:
                result_last[group]['前三页'] += num
            if page in ['首页', '二页']:
                result_last[group]['前二页'] += num
            if page == '首页':
                result_last[group]['首页'] += num
    return result_last


# 写入excel文件
def write_excel(group_list, result_last, today):
    wb = Workbook()
    # 创建sheet
    for group in group_list:
        sheet_num = 0
        wb.create_sheet(u'{0}'.format(group), index=sheet_num)
        sheet_num += 1
    # 写内容
    row_first = ['日期', '首页', '前二页', '前三页', '前四页', '前五页', '五页后']
    for group, data_dict in result_last.items():
        # 写第一行表头
        wb[group].append(row_first)
        # 写数据
        row_value = [today]
        for page, value in data_dict.items():
            row_value.append(value)
        wb[u'{0}'.format(group)].append(row_value)
    wb.save('{0}bdmo1_page5.xlsx'.format(today))


class bdmoCoverIndex(threading.Thread):

    def __init__(self):
        threading.Thread.__init__(self)

    @staticmethod
    def read_excel(filepath):
        q = queue.Queue()
        group_list = []
        kwd_dict = {}
        wb_kwd = load_workbook(filepath)
        for sheet_obj in wb_kwd:
            sheet_name = sheet_obj.title
            group_list.append(sheet_name)
            kwd_dict[sheet_name]= []
            col_a = sheet_obj['A']
            for cell in col_a:
                kwd = (cell.value)
                # 加个判断吧
                if kwd:
                    kwd_z = kwd+ '租房'
                    kwd_er = kwd + '二手房'
                    q.put([sheet_name,kwd_z])
                    q.put([sheet_name,kwd_er])
        return q, group_list

    # 初始化结果字典
    @staticmethod
    def result_init(group_list):
        pages = ['首页','二页','三页','四页','五页','五页后']
        result = {} #记录首页、第二页..第五页各多少词
        for group in group_list:
            result[group] = {}
        for group in group_list:
            for page in pages:
                result[group][page] = 0
        print("结果字典init...")
        return result

    # 初始化最后结果字典
    @staticmethod
    def result_last_init(group_list):
        result_last = {}  # 记录首页、第二页..第五页各多少词
        pages = ['首页','前二页','前三页','前四页','前五页','五页后']
        for group in group_list:
            result_last[group] = {}
        for group in group_list:
            for page in pages:
                result_last[group][page] = 0
        print("最终结果字典init...")
        return result_last

    # 获取某词serp源码

    # 获取某词的serp源码上包含排名url的div块
    def get_data_logs(self, html ,url):
        data_logs = []
        doc = pq(html)
        title = doc('title').text()
        if '- 百度' in title and 'https://m.baidu.com/s' in url:
            try:
                div_list = doc('.c-result').items()
            except Exception as e:
                print('提取div块失败', e)
            else:
                for div in div_list:
                    data_log = div.attr('data-log')
                    data_logs.append(data_log) if data_log is not None else data_logs
        else:
            print(title,'源码异常,可能反爬')
        return data_logs

    # 提取排名的真实url及排名
    def get_real_urls(self, data_logs=[]):
        real_urls = []
        for data_log in data_logs:
            # json字符串要以双引号表示
            data_log = json.loads(data_log.replace("'", '"'))
            srcid = data_log['ensrcid'] if 'ensrcid' in data_log  else 'ensrcid'
            rank_url = data_log['mu'] if 'mu' in data_log else srcid # mu不存在
            rank_url = rank_url if rank_url else srcid # mu为空
            rank = data_log['order']
            real_urls.append((rank_url,rank))
        return real_urls

    # 提取某url的域名部分
    def get_domain(self,real_url):
        domain = None
        try:
           res = urlparse(real_url)
        except Exception as e:
           print (e,real_url)
        else:
           domain = res.netloc
        return domain
        
    # 获取某词serp源码首页排名所有域名
    def get_domains(self,real_url_list):
            domain_list = [self.get_domain(real_url) for real_url in real_url_list]
            # 一个词某域名多个url有排名,计算一次
            domain_set = set(domain_list)
            domain_set = domain_set.remove(None) if None in domain_set else domain_set
            domain_str = ','.join(domain_set)
            return domain_str

    # 线程函数
    def run(self):
        driver = webdriver.Chrome(options=option,chrome_options=option)
        while 1:
                group_kwd = q.get()
                group,kwd = group_kwd
                # group,kwd = ('北京','经开区二手房')
                print(group,kwd)
                try:
                    driver.get('https://m.baidu.com/')
                    print(11111,driver.window_handles)
                    for k,v in cookie.items():
                        driver.add_cookie({'name':k,'value':v})
                    input = WebDriverWait(driver, 30).until(
                        EC.visibility_of_element_located((By.ID, "index-kw"))
                    )
                    input.click() # 先click后clear,直接send_keys容易丢失字符
                    input.clear()
                    for wd in kwd:
                        input.send_keys(wd) # send_keys和click之间间隔时间;一次性输入容易丢失字符
                    baidu = WebDriverWait(driver, 20).until(
                        EC.visibility_of_element_located((By.ID, "index-bn"))
                    )
                    baidu.click()  # 如果有弹窗覆盖会click失败
                    print(222222,driver.window_handles)
                    # exit()
                    # 等待首页元素加载
                    next_page = WebDriverWait(driver, 20).until(
                        EC.visibility_of_element_located((By.CLASS_NAME,"new-nextpage-only"))
                    )
                    driver.execute_script(js)
                    html_1 = driver.page_source
                    now_url = driver.current_url
                    data_logs = self.get_data_logs(html_1,now_url)
                    if data_logs:
                        real_url_list_rank = self.get_real_urls(data_logs)
                        real_url_list = []
                        for my_url,my_order in real_url_list_rank:
                            real_url_list.append(my_url)
                        domain_str = self.get_domains(real_url_list)
                        if domain in domain_str:
                            for my_url,my_order,my_attr in real_url_list_rank:
                                if domain in my_url:
                                    f.write('{0}\t{1}\t{2}\t{3}\n'.format(kwd,my_url,my_order,group))
                                    break # 取第一个排名url
                            f.flush()
                    print(333333,driver.window_handles)
                    # exit()
                except Exception as e:
                    print('异常',e,now_url)
                    time.sleep(10)
                finally:
                    del kwd
                    gc.collect()
                    q.task_done()


if __name__ == "__main__":
    start = time.time()
    local_time = time.localtime()
    today = time.strftime('%Y%m%d', local_time)
    iphone_ua = 'Mozilla/5.0 (iPhone; CPU iPhone OS 11_0 like Mac OS X) AppleWebKit/604.1.38 (KHTML, like Gecko) Version/11.0 Mobile/15A372 Safari/604.1'
    option = Options()
    # option.add_argument('--window-size=380,700')
    option.add_argument('--disable-gpu')
    option.add_argument("--disable-notifications")
    option.binary_location = "C:/Program Files (x86)/Google/Chrome/Application/chrome.exe"  #安装的位置
    # option.add_experimental_option("excludeSwitches", ['enable-automation']);
    option.add_argument('disable-infobars')
    mobile_emulation = {
        "deviceMetrics": {"width": 375, "height": 667, 'pixelRatio': 3},
        "userAgent": iphone_ua
    }

    prefs = {
        'profile.default_content_setting_values': {
            # 'images' : 2, # 禁止图片加载
            'notifications': 2,  # 禁止弹窗
            'profile.default_content_settings.popups': 2,
        }
    }
    cookie_str = 'wpr=0; ___rl__test__cookies=1578966798931; BDICON=10123156; BIDUPSID=95E739A8EE050812705C1FDE2584A61E; PSTM=1563865961; BAIDUID=95E739A8EE050812705C1FDE2584A61E:SL=0:NR=10:FG=1; BDUSS=NMRzZPVUFqR0JtbzJJc1ZDdkx2MGtiQUpvWVNUSjhnSUFmRFRmTnpDdmpGcXhkRVFBQUFBJCQAAAAAAAAAAAEAAADag5oxzI2IkAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAOOJhF3jiYRdOU; MSA_PBT=146; plus_lsv=f197ee21ffd230fd; plus_cv=1::m:49a3f4a6; MSA_ZOOM=1056; BDORZ=FFFB88E999055A3F8A630C64834BD6D0; wpr=1; ysm=8407|10303; SE_LAUNCH=5%3A26315185_0%3A26315185_25%3A26315185; BDRCVFR[xoix5KwSHTc]=9xWipS8B-FspA7EnHc1QhPEUf; delPer=0; SIGNIN_UC=70a2711cf1d3d9b1a82d2f87d633bd8a03289781966; H_WISE_SIDS=141144_139203_139419_139403_137831_114177_139396_135846_141000_139148_120169_139864_140833_133995_138878_137979_140173_131247_132552_140227_118880_118865_118839_118832_118793_138165_107317_138883_140260_139046_140592_139174_139624_140114_136196_131861_140591_139694_138586_133847_140792_140545_134256_131423_140823_138663_136537_141103_110085_140325_127969_140622_141131_140595_140864_139802_137252_139408_127417_138312_138425_141193_140685_141190_140596_138754_140964; rsv_i=067cYjOPPh%2BA3DtTguvjjCd8iEFSe3pHyAzI6PHKyd49PsuvYo3eu%2FlsP8rhMtSIBQaJQV70fy%2Bd1GE7dWHRnyUh%2F8iC0gw; FEED_SIDS=345657_0114_8; ___rl__test__cookies=1578966786128; wise_tj_ub=ci%4081_32_-1_-1_-1_-1_-1%7Ciq%4019_1_18_271%7Ccb%40-1_-1_-1_-1_-1_-1_-1%7Cce%401%7Ctse%401; BDICON=10123156; BDPASSGATE=IlPT2AEptyoA_yiU4V_43kIN8enzTri4H4PISkpT36ePdCyWmhHWBAREUjD6YnSgBC3gzDDPdstPoifKXlVXa_EqnBsZolpMany5xNSCgsTtPsx17QovIab2KUE2sA8PbRhL-3MJJ3NUMWosyBDxhAY1fe768Qx5huvRrzHgmMjsAkeR3oj6r7aTY767O-0APNuc-R0QbSh-OkOWVOGxRILYhFchOJ1L70aOatY6C3D5q6oY0RuiZMExGI8mFppi_x3nBQOLkKaoEV55qysc; BDSVRTM=78; BDORZ=SFH; ASUV=1.2.126; MSA_WH=414_736; COOKIE_SESSION=55631_11_4_9_2_t1_12_9_12_0_0_8_136_1578966790%7C9%230_0_0_0_0_0_0_0_1578902935%7C1; FC_MODEL=0_11_2_0_4.59_0_4_0_0_0_6.86_0_9_12_6_38_0_0_1578966790742%7C9%234.59_0_0_9_6_0_1578966790742%7C9; OUTFOX_SEARCH_USER_ID_NCOO=1661554528.3256326; BDSVRBFE=Go; __bsi=8939560779121142338_00_20_N_R_10_0303_c02f_Y; wise_tj_ub=ci%4043_23_-1_-1_-1_-1_-1%7Ciq%404_1_4_435%7Ccb%40-1_-1_-1_-1_-1_-1_-1%7Cce%401%7Ctse%401'
    lists = cookie_str.split(';')
    cookie = {}
    for i in lists:
        j = i.strip()
        j = j.split('=')
        cookie[j[0]] = j[1]
    option.add_experimental_option('prefs', prefs)
    option.add_experimental_option("mobileEmulation", mobile_emulation)
    js = 'window.scrollBy(0,{0})'.format('document.body.scrollHeight')
    domain = '5i5j.com'
    q,group_list = bdmoCoverIndex.read_excel('kwd_core_city.xlsx')  # 关键词队列及分类
    result = bdmoCoverIndex.result_init(group_list)  # 结果字典
    result_last = bdmoCoverIndex.result_last_init(group_list)  # 最终结果字典
    all_num = q.qsize() # 总词数
    f = open('{0}bdmo1_page5_serpurl.txt'.format(today),'w',encoding="utf-8")
    file_path = f.name
    # 设置线程数
    for i in list(range(1)):
        t = bdmoCoverIndex()
        t.setDaemon(True)
        t.start()
    q.join()
    f.close()
    end = time.time()
    print('耗时{0}min'.format((end-start)/60))

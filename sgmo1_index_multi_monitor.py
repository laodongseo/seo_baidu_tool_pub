# ‐*‐ coding: utf‐8 ‐*‐
"""
功能:
   1)指定几个域名,分关键词种类监控首页词数
   2)采集serp所有url,提取域名并统计各域名首页覆盖率
   3)采集了serp上的排名url特征srcid值
提示:
  1)相关网站.相关企业.智能小程序.其他人还在搜.热议聚合.资讯聚合.搜索智能聚合.视频全部算在内
    所以首页排名有可能大于10
  2)serp上自然排名mu属性值为排名url,特殊样式mu为空或不存在,
    提取article里url,该url是baidu域名,二次访问才能获得真实url,本脚本直接取baidu链接
  3)2020kwd_url_core_city_unique.xlsx:sheet名为关键词种类,sheet第一列放关键词
结果:
    sgmo1_index_info.txt:各监控站点词的排名及url,如有2个url排名,只取第一个
    sgmo1_index_all.txt:serp所有url及样式特征,依此统计各域名首页覆盖率-单写脚本完成
    sgmo1_index.xlsx:自己站每类词首页词数
    sgmo1_index_domains.xlsx:各监控站点每类词的首页词数
    sgmo1_index_domains.txt:各监控站点每类词的首页词数
header头信息复制浏览器全部的请求头,Accept-Encoding留deflate
搜狗M端源码极其不规范,必须处理后才能正常解析
"""

import requests
from pyquery import PyQuery as pq
import threading
import queue
import time
from urllib.parse import urlparse
from openpyxl import load_workbook
from openpyxl import Workbook
import time
import gc
import json
import random
import re
from urllib import request
from http import cookiejar
import urllib3


# 设置忽略SSL验证
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


# 计算最终结果
def get_result(file_path, result):
    for line in open(file_path, 'r', encoding='utf-8'):
        line = line.strip().split('\t')
        rank = line[2]
        group = line[3]
        domain = line[4]
        if rank != '无':
            result[domain][group]['首页'] += 1
        result[domain][group]['总词数'] += 1
    return result


# 写txt,所有监控域名的结果
def write_domains_txt(result_last):
    with open('{0}sgmo1_index_domains.txt'.format(today), 'w', encoding="utf-8") as f_res:
        f_res.write('{0}\t{1}\t{2}\t{3}\t{4}\n'.format('日期', '域名', '词类', '首页词数', '查询词数'))
        for now_domain, dict_value in result_last.items():
            for group, dict_index_all in dict_value.items():
                f_res.write('{0}\t{1}\t{2}\t'.format(today, now_domain, group))
                for key, value in dict_index_all.items():
                    f_res.write(str(value) + '\t')
                f_res.write('\n')


# 写excel
def write_myexcel(group_list, result_last, today, my_domain):
    wb = Workbook()
    wb_all = Workbook()
    # 创建sheet写表头
    for group in group_list:
        sheet_num = 0
        wb.create_sheet(u'{0}'.format(group), index=sheet_num)
        wb_all.create_sheet(u'{0}'.format(group), index=sheet_num)
        row_first = ['日期', '首页', '总词数']
        row_first2 = ['日期', '域名', '首页', '总词数']
        # 写表头
        wb[group].append(row_first)
        wb_all[group].append(row_first2)
        sheet_num += 1
    # 写内容
    for domain, dict_value in result_last.items():
        if domain == my_domain:
            for group, dict_index_all in dict_value.items():
                # 写数据
                row_value = [today]
                for key, value in dict_index_all.items():
                    row_value.append(value)
                wb[u'{0}'.format(group)].append(row_value)

        for group, dict_index_all in dict_value.items():
            # 写数据
            row_value = [today, domain]
            for key, value in dict_index_all.items():
                row_value.append(value)
            wb_all[u'{0}'.format(group)].append(row_value)
    wb.save('{0}sgmo1_index.xlsx'.format(today))
    wb_all.save('{0}sgmo1_index_domains.xlsx'.format(today))


# header
def get_header(COOKIE):
    headers = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
        'Accept-Encoding': 'gzip, deflate, br',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        'Connection': 'keep-alive',
        'Host': 'm.sogou.com',
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.157 Safari/537.36',
        'Cookie': COOKIE
    }
    return headers


def get_cookie(file_path):
    COOKIE = ''
    try:
        COOKIE = open(file_path).read().strip()
    except FileNotFoundError:
        cookies_file = open(file_path, 'w', encoding='utf-8')
        cookies_file.close()
    if COOKIE == '':
        # 声明一个CookieJar对象实例来保存cookie
        cookie = cookiejar.CookieJar()
        # 利用urllib.request库的HTTPCookieProcessor对象来创建cookie处理器,也就CookieHandler
        handler = request.HTTPCookieProcessor(cookie)
        # 通过CookieHandler创建opener
        opener = request.build_opener(handler)
        # 此处的open方法打开网页
        response = opener.open('https://m.sogou.com')
        # 打印cookie信息
        cookie_result = ""
        for item in cookie:
            cookie_result = cookie_result + item.name + "=" + item.value + ";"
        my_open = open(file_path, 'w', encoding='utf-8')
        print('%s' % (cookie_result), file=my_open)
        my_open.close()
    return COOKIE


class sgmoIndexMonitor(threading.Thread):
    def __init__(self):
        threading.Thread.__init__(self)

    @staticmethod
    def read_excel(filepath):
        q = queue.Queue()
        group_list = []
        kwd_dict = {}
        wb_kwd = load_workbook(filepath)
        for sheet_obj in wb_kwd:
            sheet_name = sheet_obj.title
            group_list.append(sheet_name)
            kwd_dict[sheet_name] = []
            col_a = sheet_obj['A']
            for cell in col_a:
                kwd = (cell.value)
                # 加个判断吧
                if kwd:
                    q.put([sheet_name, kwd])
        return q, group_list

    # 初始化结果字典
    @staticmethod
    def result_init(group_list):
        result = {}
        for domain in domains:
            result[domain] = {}
            for group in group_list:
                result[domain][group] = {'首页': 0, '总词数': 0}
        print("结果字典init...")
        return result

    # 获取某词serp源码
    def get_html(self, url, my_header, retry=1):
        try:
            r = requests.get(url=url, headers=my_header, timeout=15)
        except Exception as e:
            print('获取源码失败', e)
            time.sleep(30)
            if retry > 0:
                self.get_html(url, my_header, retry - 1)
        else:
            html = r.content.decode('utf-8', errors='ignore')  # 用r.text有时候识别错误
            url = r.url  # 反爬会重定向,取定向后的地址
            return html, url

    # 获取某词serp源码所有url
    def get_encrpt_urls(self, html, url):
        encrypt_url_list = []
        html = html.replace('</body>','')
        html = html.replace('</html>','')
        html = html.replace('class=results','class="results"')
        html = re.sub('\0','',html)
        html = html + '</body></html>'
        # with open('111.html','w',encoding='utf-8') as f:
        #     f.write(html)
        doc = pq(html,parser="html")
        if 'https://m.sogou.com/web/searchList.jsp' in url:
            div_contents = doc('.results .vrResult').items()
            for div in div_contents:
                data_pos = div.attr('data-pos')
                if not data_pos:
                    rank_id = div.attr('id')
                    if rank_id:
                        encrypt_url = div('h3 a').attr('href')
                        # 过滤掉xxx_相关 该样式是link标签
                        if encrypt_url and encrypt_url.startswith('javascript:void'):  # 搜狗立知
                            encrypt_url = div('h3 span a').attr('href')
                    else:
                        a = div('h3 a')
                        encrypt_url = a.attr('href')
                        rank_id = a.attr('id')
                    rank = rank_id.split('_')[-1]
                    if encrypt_url:
                        encrypt_url_list.append((encrypt_url, rank))
        else:
            print('源码异常,可能反爬')
            time.sleep(60)
        return encrypt_url_list

    # 解密某条加密url
    def decrypt_url(self, encrypt_url, my_header, retry=1):
        real_url = None  # 默认None
        if encrypt_url.startswith('http'):
            real_url = encrypt_url
        else:
            if encrypt_url.startswith('./'):
                encrypt_url = encrypt_url.replace('./', '',1)
                encrypt_url = '{0}{1}'.format('https://m.sogou.com/web/', encrypt_url)
                try:
                    r = requests.get(encrypt_url, verify=False, headers=my_header)
                except Exception as e:
                    print(encrypt_url, '获取加密地址源码失败', e)
                    myf.write(encrypt_url + '\n')
                    myf.flush()
                    time.sleep(60)
                    if retry > 0:
                        self.decrypt_url(encrypt_url, my_header, retry - 1)
                else:
                    html = r.text
                    doc = pq(html,parser="html")
                    a = doc('.btn a')
                    real_url = a.attr('href')
        return real_url


    # 提取某url的域名部分
    def get_domain(self, real_url):
        domain = None
        try:
            res = urlparse(real_url)
        except Exception as e:
            print(e, real_url)
        else:
            domain = res.netloc
        return domain

    # 获取某词serp源码首页排名所有域名
    def get_domains(self, real_url_list):
        domain_list = [self.get_domain(real_url) for real_url in real_url_list]
        # 一个词某域名多个url有排名,计算一次
        domain_set = set(domain_list)
        domain_set = domain_set.remove(None) if None in domain_set else domain_set
        domain_str = ','.join(domain_set)
        return domain_str

    # 线程函数
    def run(self):
        while 1:
            group_kwd = q.get()
            group, kwd = group_kwd
            print(group, kwd)
            try:
                url = "https://m.sogou.com/web/searchList.jsp?pid=sogou-waps-7880d7226e87b77&keyword={0}".format(kwd)
                cookie = get_cookie(cookie_txt)
                my_header = get_header(cookie)
                html, now_url = self.get_html(url, my_header)
                encrypt_url_list_rank = self.get_encrpt_urls(html, now_url)
                real_urls_rank = []
                # 源码ok再写入
                if encrypt_url_list_rank:
                    for my_serp_url, my_order in encrypt_url_list_rank:
                        my_real_url = self.decrypt_url(my_serp_url, my_header)
                        print(my_real_url)
                        real_urls_rank.append((my_real_url, my_order))
                        time.sleep(4)
                    real_urls = []  # 只存储url
                    for my_real_url, my_order in real_urls_rank:
                        real_urls.append(my_real_url)
                        f_all.write('{0}\t{1}\t{2}\t{3}\n'.format(kwd, my_real_url, my_order, group))
                    print(real_urls)
                    domain_str = self.get_domains(real_urls)
                    # 目标站点是否出现
                    for domain in domains:
                        if domain not in domain_str:
                            f.write('{0}\t{1}\t{2}\t{3}\t{4}\n'.format(kwd, '无', '无', group, domain))
                        else:
                            # my_url可能为None
                            for my_url, my_order in real_urls_rank:
                                if domain in str(my_url):
                                    f.write('{0}\t{1}\t{2}\t{3}\t{4}\n'.format(kwd, my_url, my_order, group, domain))
                                    # print(my_url, my_order)
                                    break  # 取第一个排名url
                f.flush()
            except Exception as e:
                print(e)
            finally:
                del kwd
                gc.collect()
                q.task_done()
                # time.sleep(3)


if __name__ == "__main__":
    start = time.time()
    myf = open('sfmo_fail.txt', 'w', encoding='utf-8')
    cookie_txt = 'cookies.txt'
    local_time = time.localtime()
    today = time.strftime('%Y%m%d', local_time)
    domains = ['5i5j.com', 'lianjia.com', 'anjuke.com', 'fang.com']  # 目标域名
    my_domain = '5i5j.com'  # 自己域名
    q, group_list = sgmoIndexMonitor.read_excel('2020kwd_url_core_city_unique.xlsx')  # 关键词队列及分类
    result = sgmoIndexMonitor.result_init(group_list)  # 初始化结果
    all_num = q.qsize()  # 总词数
    f = open('{0}sgmo1_index_info.txt'.format(today), 'w', encoding="utf-8")
    f_all = open('{0}sgmo1_index_all.txt'.format(today), 'w', encoding="utf-8")
    file_path = f.name
    # 设置线程数
    for i in list(range(1)):
        t = sgmoIndexMonitor()
        t.setDaemon(True)
        t.start()
    q.join()
    f.close()
    f_all.close()
    # 根据sgmo1_index_info.txt计算结果
    result_last = get_result(file_path, result)
    # 写入txt文件
    write_domains_txt(result_last)
    # 写入excel
    write_myexcel(group_list, result_last, today, my_domain)
    # 统计查询成功的词数
    with open(file_path, 'r', encoding='utf-8') as fp:
        success = int(sum(1 for x in fp) / len(domains))
    end = time.time()
    print('关键词共{0}个,耗时{1}min'.format(all_num, (end - start) / 60))
